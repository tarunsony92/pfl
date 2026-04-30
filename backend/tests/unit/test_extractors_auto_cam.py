"""Tests for AutoCamExtractor — happy path and degraded paths."""

import io
from pathlib import Path

import openpyxl
import pytest

from app.enums import ExtractionStatus
from app.worker.extractors.auto_cam import AutoCamExtractor
from tests.fixtures.builders.auto_cam_builder import build_auto_cam_xlsx

# ---------------------------------------------------------------------------
# Happy path
# ---------------------------------------------------------------------------


async def test_auto_cam_partial_when_key_field_missing(tmp_path: Path):
    """All 4 sheets present but a key field (applicant_name) is empty → PARTIAL."""
    p = build_auto_cam_xlsx(tmp_path / "cam_empty_name.xlsx", applicant_name="")
    body = p.read_bytes()
    result = await AutoCamExtractor().extract("cam_empty_name.xlsx", body)

    assert result.status == ExtractionStatus.PARTIAL
    assert result.error_message is None


async def test_auto_cam_extracts_all_fields_from_standard_fixture(tmp_path: Path):
    p = build_auto_cam_xlsx(
        tmp_path / "cam.xlsx",
        applicant_name="TEST USER",
        cibil_score=800,
        foir=0.35,
        loan_amount=200000,
    )
    body = p.read_bytes()
    result = await AutoCamExtractor().extract("cam.xlsx", body)

    assert result.status == ExtractionStatus.SUCCESS
    assert result.schema_version == "1.0"
    assert result.warnings == []
    assert result.error_message is None

    # system_cam
    sc = result.data["system_cam"]
    assert sc["applicant_name"] == "TEST USER"
    assert sc["pan"] == "ABCDE1234F"
    assert sc["loan_amount"] == 200000

    # eligibility
    el = result.data["eligibility"]
    assert el["cibil_score"] == 800
    assert el["foir"] == pytest.approx(0.35)
    assert el["eligible_amount"] == 200000

    # cm_cam_il
    cam = result.data["cm_cam_il"]
    assert cam["borrower_name"] == "TEST USER"
    assert cam["pan_number"] == "ABCDE1234F"
    assert cam["cibil"] == 800

    # health_sheet
    hs = result.data["health_sheet"]
    assert hs["total_monthly_income"] == 35000
    assert hs["total_monthly_expense"] == 14700
    assert hs["net_surplus"] == 20300


async def test_auto_cam_extractor_name_and_schema_version():
    extractor = AutoCamExtractor()
    assert extractor.extractor_name == "auto_cam"
    assert extractor.schema_version == "1.0"


# ---------------------------------------------------------------------------
# Degraded: one sheet missing → PARTIAL
# ---------------------------------------------------------------------------


async def test_auto_cam_partial_when_sheet_missing(tmp_path: Path):
    p = build_auto_cam_xlsx(tmp_path / "cam.xlsx")
    # Remove the SystemCam sheet
    wb = openpyxl.load_workbook(p)
    del wb["SystemCam"]
    wb.save(p)

    result = await AutoCamExtractor().extract("cam.xlsx", p.read_bytes())

    assert result.status == ExtractionStatus.PARTIAL
    assert "missing_sheet:SystemCam" in result.warnings
    assert result.data["system_cam"] == {}
    # Other sheets still parsed
    assert result.data["eligibility"].get("cibil_score") is not None


# ---------------------------------------------------------------------------
# Degraded: all sheets missing → FAILED
# ---------------------------------------------------------------------------


async def test_auto_cam_failed_when_no_expected_sheets(tmp_path: Path):
    """Multi-sheet workbook where none of the sheet names match any alias →
    FAILED with four missing_sheet warnings (one per required canonical sheet).
    """
    wb = openpyxl.Workbook()
    wb.active.title = "WrongSheet"  # type: ignore[union-attr]
    # Add 3 more unrelated sheets so the workbook is multi-sheet (not a
    # single-sheet variant, which is FAILED with a different code path).
    wb.create_sheet("Junk1")
    wb.create_sheet("Junk2")
    wb.create_sheet("Junk3")
    out = tmp_path / "bad.xlsx"
    wb.save(out)

    result = await AutoCamExtractor().extract("bad.xlsx", out.read_bytes())

    assert result.status == ExtractionStatus.FAILED
    assert result.error_message is not None
    assert len(result.warnings) == 4  # one per missing required sheet


# ---------------------------------------------------------------------------
# Degraded: corrupted bytes → FAILED
# ---------------------------------------------------------------------------


async def test_auto_cam_failed_on_corrupt_bytes():
    result = await AutoCamExtractor().extract("corrupt.xlsx", b"not an xlsx file")

    assert result.status == ExtractionStatus.FAILED
    assert result.error_message is not None
    assert "corrupt.xlsx" in result.error_message


# ---------------------------------------------------------------------------
# Extra: unknown labels are silently ignored
# ---------------------------------------------------------------------------


async def test_auto_cam_ignores_unknown_labels(tmp_path: Path):
    p = build_auto_cam_xlsx(tmp_path / "cam.xlsx")
    wb = openpyxl.load_workbook(p)
    ws = wb["SystemCam"]
    ws["A5"] = "Unknown Field"
    ws["B5"] = "some value"
    wb.save(p)

    result = await AutoCamExtractor().extract("cam.xlsx", p.read_bytes())
    assert result.status == ExtractionStatus.SUCCESS
    assert "unknown_field" not in result.data["system_cam"]


# ---------------------------------------------------------------------------
# Edge: label with extra whitespace / mixed case
# ---------------------------------------------------------------------------


async def test_auto_cam_normalises_label_whitespace_and_case(tmp_path: Path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet in ("SystemCam", "Elegibilty", "CM CAM IL", "Health Sheet"):
        wb.create_sheet(sheet)

    ws = wb["SystemCam"]
    ws["A1"] = "  APPLICANT  NAME  "  # extra spaces + uppercase
    ws["B1"] = "NORMALISED USER"
    ws["A2"] = "Date of Birth"
    ws["B2"] = "01/01/1990"
    ws["A3"] = "PAN"
    ws["B3"] = "ZZZZZ9999Z"
    ws["A4"] = "Loan Amount"
    ws["B4"] = 100000

    ws_e = wb["Elegibilty"]
    ws_e["A1"] = "CIBIL Score"
    ws_e["B1"] = 750
    ws_e["A2"] = "FOIR"
    ws_e["B2"] = 0.4
    ws_e["A3"] = "Eligible Amount"
    ws_e["B3"] = 100000

    ws_c = wb["CM CAM IL"]
    ws_c["A1"] = "Borrower Name"
    ws_c["B1"] = "NORMALISED USER"
    ws_c["A2"] = "PAN Number"
    ws_c["B2"] = "ZZZZZ9999Z"
    ws_c["A3"] = "Loan Required"
    ws_c["B3"] = 100000
    ws_c["A4"] = "CIBIL"
    ws_c["B4"] = 750

    ws_h = wb["Health Sheet"]
    ws_h["A1"] = "Total Monthly Income"
    ws_h["B1"] = 50000
    ws_h["A2"] = "Total Monthly Expense"
    ws_h["B2"] = 20000
    ws_h["A3"] = "Net Surplus"
    ws_h["B3"] = 30000

    buf = io.BytesIO()
    wb.save(buf)

    result = await AutoCamExtractor().extract("normalised.xlsx", buf.getvalue())
    assert result.status == ExtractionStatus.SUCCESS
    assert result.data["system_cam"]["applicant_name"] == "NORMALISED USER"


# ---------------------------------------------------------------------------
# M4 Fix 5: Fuzzy sheet-name and label matching (real-file layout)
# ---------------------------------------------------------------------------


async def test_auto_cam_fuzzy_sheet_name_alias(tmp_path: Path):
    """Sheet named 'Eligibility' (not 'Elegibilty') is matched via alias."""
    import io as _io

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    wb.create_sheet("SystemCam")
    wb.create_sheet("Eligibility")   # common alternate spelling
    wb.create_sheet("CM CAM IL")
    wb.create_sheet("Health Sheet")

    ws = wb["SystemCam"]
    ws["A1"] = "Applicant Name"
    ws["B1"] = "FUZZY USER"
    ws["A2"] = "Date of Birth"
    ws["B2"] = "01/01/1990"
    ws["A3"] = "PAN"
    ws["B3"] = "AABBB1234C"
    ws["A4"] = "Loan Amount"
    ws["B4"] = 100000

    ws_e = wb["Eligibility"]
    ws_e["A1"] = "Cibil Score"
    ws_e["B1"] = 720
    ws_e["A2"] = "FOIR"
    ws_e["B2"] = 0.38
    ws_e["A3"] = "Eligible Amount"
    ws_e["B3"] = 100000

    ws_c = wb["CM CAM IL"]
    ws_c["A1"] = "Borrower Name"
    ws_c["B1"] = "FUZZY USER"
    ws_c["A2"] = "PAN Number"
    ws_c["B2"] = "AABBB1234C"
    ws_c["A3"] = "Loan Required"
    ws_c["B3"] = 100000
    ws_c["A4"] = "CIBIL"
    ws_c["B4"] = 720

    ws_h = wb["Health Sheet"]
    ws_h["A1"] = "Total Monthly Income"
    ws_h["B1"] = 30000
    ws_h["A2"] = "Total Monthly Expense"
    ws_h["B2"] = 12000
    ws_h["A3"] = "Net Surplus"
    ws_h["B3"] = 18000

    buf = _io.BytesIO()
    wb.save(buf)

    result = await AutoCamExtractor().extract("fuzzy_elig.xlsx", buf.getvalue())
    assert result.status == ExtractionStatus.SUCCESS
    assert result.data["eligibility"]["cibil_score"] == 720


async def test_auto_cam_real_layout_col_b_label_col_c_value(tmp_path: Path):
    """Real CAM layout: col A = section heading, col B = label, col C = value."""
    import io as _io

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Real file uses "SystemCam" for sheet name but data is in B/C columns
    ws_sys = wb.create_sheet("SystemCam")
    ws_sys["A1"] = "Personal Details"
    ws_sys["B2"] = "First Name"
    ws_sys["C2"] = "AJAY SINGH"
    ws_sys["B3"] = "Date Of Birth"
    ws_sys["C3"] = "17-11-2001"
    ws_sys["B4"] = "PAN"
    ws_sys["C4"] = "OWLPS6441C"
    ws_sys["A5"] = "Product Details"
    ws_sys["B6"] = "Loan Amount"
    ws_sys["C6"] = "100000"

    ws_elig = wb.create_sheet("Elegibilty")
    ws_elig["B1"] = "Name of Applicant"
    ws_elig["C1"] = "AJAY SINGH"
    ws_elig["B2"] = "Cibil Score"
    ws_elig["C2"] = 750
    ws_elig["B3"] = "FOIR"
    ws_elig["C3"] = 0.181

    ws_cam = wb.create_sheet("CM CAM IL")
    ws_cam["A1"] = "Profile Screening"
    ws_cam["A2"] = "Name"
    ws_cam["B2"] = "AJAY SINGH"   # label in col A, value in col B (A+B pair)
    ws_cam["A3"] = "FOIR"
    ws_cam["B3"] = 0.181
    ws_cam["A4"] = "Borrower Name"  # standard label too as fallback
    ws_cam["B4"] = "AJAY SINGH"

    ws_health = wb.create_sheet("Health Sheet")
    ws_health["A1"] = "Category"
    ws_health["B1"] = "Total Monthly Income"
    ws_health["C1"] = 36000
    ws_health["B2"] = "Total Monthly Expense"
    ws_health["C2"] = 10000

    buf = _io.BytesIO()
    wb.save(buf)

    result = await AutoCamExtractor().extract("real_layout.xlsx", buf.getvalue())
    # Should at minimum be PARTIAL (key fields may or may not all be found)
    assert result.status in (ExtractionStatus.SUCCESS, ExtractionStatus.PARTIAL)
    # SystemCam: First Name → applicant_name, Loan Amount → loan_amount
    assert result.data["system_cam"].get("applicant_name") == "AJAY SINGH"
    # Eligibility: Cibil Score → cibil_score
    assert result.data["eligibility"].get("cibil_score") == 750
    # Health sheet: Total Monthly Income from B+C pair
    assert result.data["health_sheet"].get("total_monthly_income") == 36000


async def test_auto_cam_a_col_label_c_col_value_b_empty(tmp_path: Path):
    """Real SystemCam layout: col A holds the label, col B is empty, col C is value."""
    import io as _io

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for s in ("SystemCam", "Elegibilty", "CM CAM IL", "Health Sheet"):
        wb.create_sheet(s)

    ws = wb["SystemCam"]
    ws["A1"] = "First Name"
    ws["C1"] = "AJAY SINGH"  # B is empty
    ws["A2"] = "Date Of Birth"
    ws["C2"] = "17-11-2001"
    ws["A3"] = "PAN"
    ws["C3"] = "OWLPS6441C"
    ws["A4"] = "Loan Amount"
    ws["C4"] = 100000

    # Give other sheets minimal content so they exist; only SystemCam drives the test.
    wb["CM CAM IL"]["A1"] = "Borrower Name"
    wb["CM CAM IL"]["B1"] = "AJAY SINGH"

    buf = _io.BytesIO()
    wb.save(buf)
    result = await AutoCamExtractor().extract("a_c_layout.xlsx", buf.getvalue())

    # A-label / C-value (B empty) must be extracted
    assert result.data["system_cam"]["applicant_name"] == "AJAY SINGH"
    assert result.data["system_cam"]["pan"] == "OWLPS6441C"
    assert result.data["system_cam"]["loan_amount"] == 100000
    # All 4 sheets present + primary (applicant_name + pan) found → SUCCESS
    assert result.status == ExtractionStatus.SUCCESS


async def test_auto_cam_success_when_cibil_blank_but_primary_present(tmp_path: Path):
    """Blank CIBIL must NOT flip SUCCESS → PARTIAL: CIBIL is not a primary identifier."""
    p = build_auto_cam_xlsx(tmp_path / "cam_blank_cibil.xlsx")
    # Wipe the CIBIL value so only the label remains
    wb = openpyxl.load_workbook(p)
    wb["Elegibilty"]["B1"] = None
    wb["CM CAM IL"]["B4"] = None
    wb.save(p)

    result = await AutoCamExtractor().extract("cam_blank_cibil.xlsx", p.read_bytes())
    assert result.status == ExtractionStatus.SUCCESS, (
        f"expected SUCCESS with blank CIBIL, got {result.status} warnings={result.warnings}"
    )


async def test_auto_cam_cam_report_single_sheet_variant_is_success(tmp_path: Path):
    """Single-sheet 'CAM_REPORT' file is a designed variant, not a partial
    full-CAM. Three canonical sheets are absent BY DESIGN, so missing_sheet
    warnings must be suppressed and status must be SUCCESS when the single
    sheet delivers the primary output.
    """
    import io as _io

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("CAM_REPORT")
    ws["A1"] = "Credit Assessment Report"
    ws["B2"] = "First Name"
    ws["C2"] = "AJAY SINGH"
    ws["B3"] = "Date Of Birth"
    ws["C3"] = "17-11-2001"
    ws["B4"] = "PAN"
    ws["C4"] = "OWLPS6441C"

    buf = _io.BytesIO()
    wb.save(buf)

    result = await AutoCamExtractor().extract("cam_report.xlsx", buf.getvalue())
    assert result.status == ExtractionStatus.SUCCESS
    # NO missing_sheet warnings for a single-sheet variant
    missing = [w for w in result.warnings if w.startswith("missing_sheet:")]
    assert missing == []
    # Variant flag surfaced so downstream knows this is the single-sheet form
    assert result.data.get("variant") == "single_sheet_cam"
    assert result.data["system_cam"].get("applicant_name") == "AJAY SINGH"


async def test_auto_cam_single_sheet_unrecognised_fails(tmp_path: Path):
    """Single-sheet workbook whose only sheet matches no alias → FAILED."""
    import io as _io

    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("RandomSheet")
    ws["A1"] = "Irrelevant"
    buf = _io.BytesIO()
    wb.save(buf)

    result = await AutoCamExtractor().extract("random.xlsx", buf.getvalue())
    assert result.status == ExtractionStatus.FAILED
