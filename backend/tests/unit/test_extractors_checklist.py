"""Tests for ChecklistExtractor — happy path and degraded paths."""

from pathlib import Path

import openpyxl

from app.enums import ExtractionStatus
from app.worker.extractors.checklist import ChecklistExtractor
from tests.fixtures.builders.checklist_builder import build_checklist_xlsx

# ---------------------------------------------------------------------------
# Happy path
# ---------------------------------------------------------------------------


async def test_checklist_extracts_sections_and_counts(tmp_path: Path):
    p = build_checklist_xlsx(
        tmp_path / "checklist.xlsx",
        yes_keys=["Aadhaar Card", "PAN Card"],
        no_keys=["Latest Salary Slip"],
        na_keys=["Passport Photo"],
    )
    body = p.read_bytes()
    result = await ChecklistExtractor().extract("checklist.xlsx", body)

    assert result.status == ExtractionStatus.SUCCESS
    assert result.schema_version == "1.0"
    assert result.warnings == []
    assert result.error_message is None

    data = result.data
    assert data["total_items"] == 9  # 3 sections × 3 items each
    assert data["yes_count"] == 2
    assert data["no_count"] == 1
    assert data["na_count"] == 6  # 4 explicit NA + remainder

    # Check specific section/item structure
    kyc = data["sections"]["KYC Documents"]
    assert kyc["Aadhaar Card"]["value"] == "YES"
    assert kyc["PAN Card"]["value"] == "YES"
    assert kyc["Passport Photo"]["value"] == "NA"


async def test_checklist_extractor_name_and_schema_version():
    extractor = ChecklistExtractor()
    assert extractor.extractor_name == "checklist"
    assert extractor.schema_version == "1.0"


async def test_checklist_default_mode_all_yes(tmp_path: Path):
    """Builder with no yes_keys/no_keys defaults all items to Yes."""
    p = build_checklist_xlsx(tmp_path / "checklist.xlsx")
    result = await ChecklistExtractor().extract("checklist.xlsx", p.read_bytes())

    assert result.status == ExtractionStatus.SUCCESS
    assert result.data["yes_count"] == 9
    assert result.data["no_count"] == 0
    assert result.data["na_count"] == 0


# ---------------------------------------------------------------------------
# Degraded: empty xlsx → FAILED
# ---------------------------------------------------------------------------


async def test_checklist_failed_on_empty_sheet(tmp_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Empty"  # type: ignore[union-attr]
    out = tmp_path / "empty.xlsx"
    wb.save(out)

    result = await ChecklistExtractor().extract("empty.xlsx", out.read_bytes())

    assert result.status == ExtractionStatus.FAILED
    assert result.error_message is not None
    assert result.data == {}


# ---------------------------------------------------------------------------
# Degraded: header-only sheet (no items) → FAILED
# ---------------------------------------------------------------------------


async def test_checklist_failed_when_only_header_row(tmp_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Section"  # type: ignore[union-attr]
    ws["B1"] = "Item"  # type: ignore[union-attr]
    ws["C1"] = "Status"  # type: ignore[union-attr]
    out = tmp_path / "header_only.xlsx"
    wb.save(out)

    result = await ChecklistExtractor().extract("header_only.xlsx", out.read_bytes())

    assert result.status == ExtractionStatus.FAILED
    assert result.error_message is not None


# ---------------------------------------------------------------------------
# Degraded: very few items → PARTIAL
# ---------------------------------------------------------------------------


async def test_checklist_partial_when_fewer_than_3_items(tmp_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Section"  # type: ignore[union-attr]
    ws["B1"] = "Item"  # type: ignore[union-attr]
    ws["C1"] = "Status"  # type: ignore[union-attr]
    ws["A2"] = "KYC Documents"  # type: ignore[union-attr]  # section header
    ws["A3"] = "KYC Documents"  # type: ignore[union-attr]
    ws["B3"] = "Aadhaar Card"  # type: ignore[union-attr]
    ws["C3"] = "Yes"  # type: ignore[union-attr]
    ws["A4"] = "KYC Documents"  # type: ignore[union-attr]
    ws["B4"] = "PAN Card"  # type: ignore[union-attr]
    ws["C4"] = "No"  # type: ignore[union-attr]
    out = tmp_path / "few.xlsx"
    wb.save(out)

    result = await ChecklistExtractor().extract("few.xlsx", out.read_bytes())

    assert result.status == ExtractionStatus.PARTIAL
    assert "very_few_items" in result.warnings
    assert result.data["total_items"] == 2


# ---------------------------------------------------------------------------
# Degraded: corrupted bytes → FAILED
# ---------------------------------------------------------------------------


async def test_checklist_failed_on_corrupt_bytes():
    result = await ChecklistExtractor().extract("bad.xlsx", b"garbage data")

    assert result.status == ExtractionStatus.FAILED
    assert result.error_message is not None
    assert "bad.xlsx" in result.error_message


# ---------------------------------------------------------------------------
# Edge: remarks column present
# ---------------------------------------------------------------------------


async def test_checklist_captures_remarks(tmp_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Section"  # type: ignore[union-attr]
    ws["B1"] = "Item"  # type: ignore[union-attr]
    ws["C1"] = "Status"  # type: ignore[union-attr]
    ws["D1"] = "Remarks"  # type: ignore[union-attr]
    ws["A2"] = "KYC"  # type: ignore[union-attr]  # section header (B2 empty)
    for i, (item, status, remark) in enumerate(
        [
            ("Aadhaar", "Yes", "Original seen"),
            ("PAN", "No", "Expired"),
            ("Photo", "NA", None),
        ],
        start=3,
    ):
        ws.cell(row=i, column=1, value="KYC")
        ws.cell(row=i, column=2, value=item)
        ws.cell(row=i, column=3, value=status)
        ws.cell(row=i, column=4, value=remark)
    out = tmp_path / "remarks.xlsx"
    wb.save(out)

    result = await ChecklistExtractor().extract("remarks.xlsx", out.read_bytes())

    assert result.status == ExtractionStatus.SUCCESS
    kyc = result.data["sections"]["KYC"]
    assert kyc["Aadhaar"]["remarks"] == "Original seen"
    assert kyc["PAN"]["remarks"] == "Expired"
    assert kyc["Photo"]["remarks"] is None


# ---------------------------------------------------------------------------
# Edge: status normalised to uppercase
# ---------------------------------------------------------------------------


async def test_checklist_normalises_status_case(tmp_path: Path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws["A1"] = "Section"  # type: ignore[union-attr]
    ws["B1"] = "Item"  # type: ignore[union-attr]
    ws["C1"] = "Status"  # type: ignore[union-attr]
    ws["A2"] = "Docs"  # type: ignore[union-attr]
    for i, (item, status) in enumerate(
        [("Item1", "yes"), ("Item2", "No"), ("Item3", "nA")],
        start=3,
    ):
        ws.cell(row=i, column=1, value="Docs")
        ws.cell(row=i, column=2, value=item)
        ws.cell(row=i, column=3, value=status)
    out = tmp_path / "case.xlsx"
    wb.save(out)

    result = await ChecklistExtractor().extract("case.xlsx", out.read_bytes())

    assert result.data["yes_count"] == 1
    assert result.data["no_count"] == 1
    assert result.data["na_count"] == 1
    docs = result.data["sections"]["Docs"]
    assert docs["Item1"]["value"] == "YES"
    assert docs["Item2"]["value"] == "NO"
    assert docs["Item3"]["value"] == "NA"
