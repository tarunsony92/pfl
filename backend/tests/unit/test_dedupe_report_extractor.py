"""DedupeReportExtractor — parses Customer_Dedupe xlsx into structured data."""
from __future__ import annotations

import openpyxl
import pytest

from app.enums import ExtractionStatus
from app.worker.extractors.dedupe_report import DedupeReportExtractor
from tests.fixtures.builders.dedupe_builder import build_dedupe_xlsx


@pytest.mark.asyncio
async def test_blank_xlsx_yields_zero_rows(tmp_path):
    """Header-only workbook (customers=[]) -> row_count == 0."""
    body = build_dedupe_xlsx(tmp_path / "dedupe.xlsx", customers=[]).read_bytes()
    res = await DedupeReportExtractor().extract("Customer_Dedupe.xlsx", body)
    assert res.status == ExtractionStatus.SUCCESS
    assert res.data["row_count"] == 0
    assert res.data["matched_rows"] == []
    assert res.data["matched_fields"] == []


@pytest.mark.asyncio
async def test_with_rows_xlsx_counts_rows(tmp_path):
    """Two data rows are counted and surfaced via matched_rows."""
    customers = [
        {"Customer Name": "Test One", "Aadhaar": "XXXX-1111", "PAN": "AAAAA1111A"},
        {"Customer Name": "Test Two", "Aadhaar": "XXXX-2222", "PAN": "BBBBB2222B"},
    ]
    body = build_dedupe_xlsx(
        tmp_path / "dedupe.xlsx", customers=customers
    ).read_bytes()
    res = await DedupeReportExtractor().extract("Customer_Dedupe.xlsx", body)
    assert res.status == ExtractionStatus.SUCCESS
    assert res.data["row_count"] == 2
    rows = res.data["matched_rows"]
    assert len(rows) == 2
    assert rows[0]["customer_name"] == "Test One"
    assert rows[1]["customer_name"] == "Test Two"
    # matched_fields lists keys with at least one non-null value
    assert "customer_name" in res.data["matched_fields"]
    assert "aadhaar" in res.data["matched_fields"]


@pytest.mark.asyncio
async def test_finpage_title_row_skipped(tmp_path):
    """Real Finpage shape: row 1 = 'Finpage - Home' title, row 2 = header,
    no data rows. Extractor must locate header on row 2 and report 0 rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customer_Dedupe"
    ws.cell(row=1, column=1, value="Finpage - Home")
    ws.cell(row=2, column=1, value="Customer Id")
    ws.cell(row=2, column=2, value="Full Name")
    ws.cell(row=2, column=3, value="Aadhaar Id")
    path = tmp_path / "finpage_dedupe.xlsx"
    wb.save(path)
    body = path.read_bytes()

    res = await DedupeReportExtractor().extract("Customer_Dedupe.xlsx", body)
    assert res.status == ExtractionStatus.SUCCESS
    assert res.data["row_count"] == 0
    assert res.data["matched_fields"] == []


@pytest.mark.asyncio
async def test_finpage_shape_with_one_match(tmp_path):
    """Finpage 16-col shape with one matched data row -> row_count == 1
    and the data dict uses customer_id key from the Finpage column name."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customer_Dedupe"
    ws.cell(row=1, column=1, value="Finpage - Home")
    headers = [
        "Customer Id", "Type", "Full Name", "Aadhaar Id",
        "Voter Card", "Pan Card", "DOB", "Mobile No",
    ]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=2, column=col_idx, value=h)
    # One data row
    values = ["C001", "REGULAR", "Match Person", "1111-2222-3333",
              "VOTER001", "PANXX1111X", "01/01/1990", "9999999999"]
    for col_idx, v in enumerate(values, start=1):
        ws.cell(row=3, column=col_idx, value=v)
    path = tmp_path / "finpage_match.xlsx"
    wb.save(path)
    body = path.read_bytes()

    res = await DedupeReportExtractor().extract("Customer_Dedupe.xlsx", body)
    assert res.status == ExtractionStatus.SUCCESS
    assert res.data["row_count"] == 1
    row = res.data["matched_rows"][0]
    assert row["customer_id"] == "C001"
    assert row["full_name"] == "Match Person"
    assert row["aadhaar_id"] == "1111-2222-3333"


@pytest.mark.asyncio
async def test_corrupt_xlsx_returns_failed():
    """Garbage bytes -> status FAILED with error_message populated."""
    res = await DedupeReportExtractor().extract("bad.xlsx", b"this is not an xlsx")
    assert res.status == ExtractionStatus.FAILED
    assert res.error_message is not None


@pytest.mark.asyncio
async def test_no_header_row_returns_failed(tmp_path):
    """A workbook with no 'Customer ...' row in first 5 rows -> FAILED."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "RandomData"
    ws.cell(row=1, column=1, value="Header A")
    ws.cell(row=1, column=2, value="Header B")
    ws.cell(row=2, column=1, value="data1")
    ws.cell(row=2, column=2, value="data2")
    path = tmp_path / "no_header.xlsx"
    wb.save(path)
    body = path.read_bytes()
    res = await DedupeReportExtractor().extract("export.xlsx", body)
    assert res.status == ExtractionStatus.FAILED
    assert res.error_message is not None
    assert "header" in res.error_message.lower()


def test_extractor_name_constant():
    """Pin the extractor_name token — Task 4's orchestrator queries
    case_extractions by extractor_name == 'dedupe_report'."""
    assert DedupeReportExtractor.extractor_name == "dedupe_report"


@pytest.mark.asyncio
async def test_normalises_punctuated_headers(tmp_path):
    """Headers with #, parens, slashes get cleanly normalised."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customer_Dedupe"
    headers = ["Customer Id", "Aadhaar #", "Voter (Card)", "Pan/Card"]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=h)
    ws.cell(row=2, column=1, value="C001")
    ws.cell(row=2, column=2, value="111122223333")
    ws.cell(row=2, column=3, value="VOTER001")
    ws.cell(row=2, column=4, value="PANXX1111X")
    path = tmp_path / "punctuated.xlsx"
    wb.save(path)
    body = path.read_bytes()

    res = await DedupeReportExtractor().extract("Customer_Dedupe.xlsx", body)
    assert res.status == ExtractionStatus.SUCCESS
    assert res.data["row_count"] == 1
    row = res.data["matched_rows"][0]
    # Keys are clean snake_case — no #, parens, or slashes
    assert "aadhaar" in row
    assert "voter_card" in row
    assert "pan_card" in row
    assert "customer_id" in row


@pytest.mark.asyncio
async def test_data_row_kept_when_first_column_empty(tmp_path):
    """Rows whose first column is empty are kept as long as some
    other cell has data. Protects against nullable lead columns."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customer_Dedupe"
    ws.cell(row=1, column=1, value="Type")
    ws.cell(row=1, column=2, value="Customer Id")
    ws.cell(row=1, column=3, value="Full Name")
    # First column blank, but row clearly has real data
    ws.cell(row=2, column=2, value="C001")
    ws.cell(row=2, column=3, value="Real Person")
    # Trailing all-blank row should still be skipped
    path = tmp_path / "nullable_first.xlsx"
    wb.save(path)
    body = path.read_bytes()

    res = await DedupeReportExtractor().extract("Customer_Dedupe.xlsx", body)
    assert res.status == ExtractionStatus.SUCCESS
    assert res.data["row_count"] == 1
    assert res.data["matched_rows"][0]["customer_id"] == "C001"


@pytest.mark.asyncio
async def test_metadata_row_does_not_qualify_as_header(tmp_path):
    """A row with 'Customer reviewed at' but only 2 cells must NOT be
    treated as a header — the real 3+ column header should win."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customer_Dedupe"
    ws.cell(row=1, column=1, value="Finpage - Home")
    ws.cell(row=2, column=1, value="Customer reviewed at")
    ws.cell(row=2, column=2, value="2026-01-01")
    # Real header on row 3
    headers = ["Customer Id", "Full Name", "Aadhaar Id"]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=3, column=col_idx, value=h)
    ws.cell(row=4, column=1, value="C001")
    ws.cell(row=4, column=2, value="Real Person")
    ws.cell(row=4, column=3, value="1111-2222-3333")
    path = tmp_path / "metadata_row.xlsx"
    wb.save(path)
    body = path.read_bytes()

    res = await DedupeReportExtractor().extract("Customer_Dedupe.xlsx", body)
    assert res.status == ExtractionStatus.SUCCESS
    assert res.data["row_count"] == 1
    row = res.data["matched_rows"][0]
    # Real header keys, not metadata-row keys
    assert "customer_id" in row
    assert "full_name" in row
    assert row["customer_id"] == "C001"
