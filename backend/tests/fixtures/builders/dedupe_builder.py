"""Builder for Customer Dedupe xlsx fixture."""

from pathlib import Path

import openpyxl

_DEFAULT_CUSTOMERS = [
    {
        "Customer Name": "SEEMA DEVI",
        "Aadhaar": "XXXX-XXXX-1234",
        "PAN": "ABCDE1234F",
        "Mobile": "9876543210",
        "DOB": "15/03/1985",
        "Address": "123 MG Road, Delhi - 110001",
    },
    {
        "Customer Name": "RAMESH KUMAR",
        "Aadhaar": "XXXX-XXXX-5678",
        "PAN": "FGHIJ5678K",
        "Mobile": "9123456789",
        "DOB": "22/07/1978",
        "Address": "45 Lajpat Nagar, Delhi - 110024",
    },
]

_COLUMNS = ["Customer Name", "Aadhaar", "PAN", "Mobile", "DOB", "Address"]


def build_dedupe_xlsx(
    path: Path,
    customers: list[dict] | None = None,
) -> Path:
    """Produce xlsx matching the Customer_Dedupe format.

    Columns: Customer Name, Aadhaar, PAN, Mobile, DOB, Address.
    Returns path.
    """
    if customers is None:
        customers = _DEFAULT_CUSTOMERS

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customer_Dedupe"

    # Header row
    for col_idx, col_name in enumerate(_COLUMNS, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Data rows
    for row_idx, customer in enumerate(customers, start=2):
        for col_idx, col_name in enumerate(_COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=customer.get(col_name, ""))

    wb.save(path)
    return path
