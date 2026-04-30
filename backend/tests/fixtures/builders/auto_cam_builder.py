"""Builder for Auto CAM xlsx fixture."""

from pathlib import Path

import openpyxl

_DEFAULTS = {
    "applicant_name": "SEEMA DEVI",
    "dob": "15/03/1985",
    "pan": "ABCDE1234F",
    "loan_amount": 150000,
    "cibil_score": 769,
    "foir": 0.42,
    "health_total_income": 35000,
    "health_total_expense": 14700,
}


def build_auto_cam_xlsx(path: Path, **overrides: object) -> Path:
    """Create a minimal Auto CAM xlsx with 4 expected sheets.

    Sheets: SystemCam, Elegibilty, CM CAM IL, Health Sheet.
    Key cells populated from defaults merged with overrides.
    Returns path.
    """
    cfg = {**_DEFAULTS, **overrides}

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default Sheet

    # --- SystemCam ---
    ws_sys = wb.create_sheet("SystemCam")
    ws_sys["A1"] = "Applicant Name"
    ws_sys["B1"] = cfg["applicant_name"]
    ws_sys["A2"] = "Date of Birth"
    ws_sys["B2"] = cfg["dob"]
    ws_sys["A3"] = "PAN"
    ws_sys["B3"] = cfg["pan"]
    ws_sys["A4"] = "Loan Amount"
    ws_sys["B4"] = cfg["loan_amount"]

    # --- Elegibilty ---
    ws_elig = wb.create_sheet("Elegibilty")
    ws_elig["A1"] = "CIBIL Score"
    ws_elig["B1"] = cfg["cibil_score"]
    ws_elig["A2"] = "FOIR"
    ws_elig["B2"] = cfg["foir"]
    ws_elig["A3"] = "Eligible Amount"
    ws_elig["B3"] = cfg["loan_amount"]

    # --- CM CAM IL ---
    ws_cam = wb.create_sheet("CM CAM IL")
    ws_cam["A1"] = "Borrower Name"
    ws_cam["B1"] = cfg["applicant_name"]
    ws_cam["A2"] = "PAN Number"
    ws_cam["B2"] = cfg["pan"]
    ws_cam["A3"] = "Loan Required"
    ws_cam["B3"] = cfg["loan_amount"]
    ws_cam["A4"] = "CIBIL"
    ws_cam["B4"] = cfg["cibil_score"]

    # --- Health Sheet ---
    ws_health = wb.create_sheet("Health Sheet")
    ws_health["A1"] = "Total Monthly Income"
    ws_health["B1"] = cfg["health_total_income"]
    ws_health["A2"] = "Total Monthly Expense"
    ws_health["B2"] = cfg["health_total_expense"]
    ws_health["A3"] = "Net Surplus"
    ws_health["B3"] = int(cfg["health_total_income"]) - int(cfg["health_total_expense"])

    wb.save(path)
    return path
