"""Builder for Checklist xlsx fixture."""

from pathlib import Path

import openpyxl

_DEFAULT_SECTIONS = {
    "KYC Documents": [
        "Aadhaar Card",
        "PAN Card",
        "Passport Photo",
    ],
    "Income Documents": [
        "Latest Salary Slip",
        "Bank Statement 6 months",
        "ITR Last 2 Years",
    ],
    "Property Documents": [
        "Sale Deed",
        "Property Tax Receipt",
        "NOC from Society",
    ],
}


def build_checklist_xlsx(
    path: Path,
    yes_keys: list[str] | None = None,
    no_keys: list[str] | None = None,
    na_keys: list[str] | None = None,
) -> Path:
    """Create a one-sheet checklist xlsx with section headers + item rows.

    Items in ``yes_keys`` get "Yes", ``no_keys`` get "No", ``na_keys`` get "NA".
    If ``yes_keys`` is provided, items not listed in any of the three sets get "NA"
    (explicit allow-list mode). If ``yes_keys`` is ``None``, items not in
    ``no_keys``/``na_keys`` default to "Yes" (backwards-compatible convenience).
    Returns path.
    """
    explicit_yes = yes_keys is not None
    yes_keys = yes_keys or []
    no_keys = no_keys or []
    na_keys = na_keys or []

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Checklist"

    ws["A1"] = "Section"
    ws["B1"] = "Item"
    ws["C1"] = "Status"

    row = 2
    for section, items in _DEFAULT_SECTIONS.items():
        ws.cell(row=row, column=1, value=section)
        ws.cell(row=row, column=2, value="")
        ws.cell(row=row, column=3, value="")
        row += 1
        for item in items:
            if item in no_keys:
                status = "No"
            elif item in na_keys:
                status = "NA"
            elif item in yes_keys:
                status = "Yes"
            else:
                status = "NA" if explicit_yes else "Yes"
            ws.cell(row=row, column=1, value=section)
            ws.cell(row=row, column=2, value=item)
            ws.cell(row=row, column=3, value=status)
            row += 1

    wb.save(path)
    return path
