"""Integration tests for the CAM discrepancy router + Phase 1 gate."""

from __future__ import annotations

import uuid as _uuid
from datetime import UTC, datetime

import pytest

from app.core.security import create_access_token
from app.enums import (
    CaseStage,
    ExtractionStatus,
    UserRole,
)
from app.models.case import Case
from app.models.case_extraction import CaseExtraction
from app.services import users as users_svc
from app.services.storage import StorageService, reset_storage_for_tests


@pytest.fixture(autouse=True)
def _reset_singletons():
    reset_storage_for_tests()
    yield
    reset_storage_for_tests()


@pytest.fixture
async def storage(mock_aws_services):
    import app.services.storage as _st_mod

    s = StorageService(
        region="ap-south-1",
        endpoint_url=None,
        access_key="test",
        secret_key="test",
        bucket="pfl-cases-dev",
    )
    await s.ensure_bucket_exists()
    _st_mod._instance = s
    yield s
    reset_storage_for_tests()


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


async def _token(db, email: str, role: UserRole) -> tuple[str, str]:
    user = await users_svc.create_user(
        db,
        email=email,
        password="Pass123!",
        full_name="T",
        role=role,
    )
    await db.commit()
    return str(user.id), create_access_token(subject=str(user.id))


async def _make_case(client, headers: dict) -> str:
    loan_id = f"CD-{str(_uuid.uuid4())[:8].upper()}"
    r = await client.post("/cases/initiate", headers=headers, json={"loan_id": loan_id})
    assert r.status_code == 201, r.text
    return r.json()["case_id"]


async def _seed_autocam_extraction(
    db,
    case_id: str,
    *,
    system_cam: dict,
    cm_cam_il: dict,
    eligibility: dict | None = None,
    health_sheet: dict | None = None,
) -> None:
    """Write an auto_cam extraction row directly so the detector has data."""
    ext = CaseExtraction(
        case_id=_uuid.UUID(case_id),
        artifact_id=None,
        extractor_name="auto_cam",
        schema_version="1.0",
        status=ExtractionStatus.SUCCESS,
        data={
            "system_cam": system_cam,
            "cm_cam_il": cm_cam_il,
            "eligibility": eligibility or {},
            "health_sheet": health_sheet or {},
        },
        warnings=None,
        error_message=None,
        extracted_at=datetime.now(UTC),
    )
    db.add(ext)
    await db.commit()


async def _force_stage(db, case_id: str, stage: CaseStage) -> None:
    case = await db.get(Case, _uuid.UUID(case_id))
    case.current_stage = stage
    await db.commit()


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


async def test_list_empty_when_no_extraction(client, db, storage):
    _, tok = await _token(db, "disc1@pfl.com", UserRole.AI_ANALYSER)
    hdrs = {"Authorization": f"Bearer {tok}"}
    case_id = await _make_case(client, hdrs)

    r = await client.get(f"/cases/{case_id}/cam-discrepancies", headers=hdrs)
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["total"] == 0
    assert body["unresolved_critical"] == 0
    assert body["phase1_blocked"] is False


async def test_list_flags_discrepancies(client, db, storage):
    _, tok = await _token(db, "disc2@pfl.com", UserRole.AI_ANALYSER)
    hdrs = {"Authorization": f"Bearer {tok}"}
    case_id = await _make_case(client, hdrs)
    await _seed_autocam_extraction(
        db,
        case_id,
        system_cam={"applicant_name": "AJAY SINGH", "pan": "OWLPS6441C", "foir_overall": 25.35},
        cm_cam_il={"borrower_name": "AJAY KUMAR", "pan_number": "ABCDE1234F", "foir": 0.181},
    )

    r = await client.get(f"/cases/{case_id}/cam-discrepancies", headers=hdrs)
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["total"] == 3  # name, pan, foir
    assert body["unresolved_critical"] == 2
    assert body["unresolved_warning"] == 1
    assert body["phase1_blocked"] is True
    keys = {v["field_key"] for v in body["views"]}
    assert {"applicant_name", "pan", "foir"} <= keys


async def test_resolve_justified(client, db, storage):
    _, tok = await _token(db, "disc3@pfl.com", UserRole.AI_ANALYSER)
    hdrs = {"Authorization": f"Bearer {tok}"}
    case_id = await _make_case(client, hdrs)
    await _seed_autocam_extraction(
        db,
        case_id,
        system_cam={"foir_overall": 25.35},
        cm_cam_il={"foir": 0.181},
    )

    r = await client.post(
        f"/cases/{case_id}/cam-discrepancies/foir/resolve",
        headers=hdrs,
        json={
            "kind": "JUSTIFIED",
            "comment": "Post-MD review: manual FOIR reflects adjusted obligations.",
        },
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["kind"] == "JUSTIFIED"
    assert body["corrected_value"] is None

    # Re-list: the flag remains (tolerance still breached) but is now resolved,
    # and phase1_blocked should reflect only CRITICAL unresolved (none here).
    r2 = await client.get(f"/cases/{case_id}/cam-discrepancies", headers=hdrs)
    s = r2.json()
    assert s["phase1_blocked"] is False
    foir_view = next(v for v in s["views"] if v["field_key"] == "foir")
    assert foir_view["resolution"] is not None


async def test_resolve_corrected_requires_value(client, db, storage):
    _, tok = await _token(db, "disc4@pfl.com", UserRole.AI_ANALYSER)
    hdrs = {"Authorization": f"Bearer {tok}"}
    case_id = await _make_case(client, hdrs)
    await _seed_autocam_extraction(
        db,
        case_id,
        system_cam={"pan": "OWLPS6441C"},
        cm_cam_il={"pan_number": "ABCDE1234F"},
    )

    r = await client.post(
        f"/cases/{case_id}/cam-discrepancies/pan/resolve",
        headers=hdrs,
        json={"kind": "CORRECTED_CM_IL", "comment": "Typo in manual entry."},
    )
    assert r.status_code == 400


async def test_resolve_corrected_success_and_no_edit_request(client, db, storage):
    _, tok = await _token(db, "disc5@pfl.com", UserRole.AI_ANALYSER)
    hdrs = {"Authorization": f"Bearer {tok}"}
    case_id = await _make_case(client, hdrs)
    await _seed_autocam_extraction(
        db,
        case_id,
        system_cam={"pan": "OWLPS6441C"},
        cm_cam_il={"pan_number": "ABCDE1234F"},
    )

    r = await client.post(
        f"/cases/{case_id}/cam-discrepancies/pan/resolve",
        headers=hdrs,
        json={
            "kind": "CORRECTED_CM_IL",
            "comment": "Manual entry typo — correcting to finpage value.",
            "corrected_value": "OWLPS6441C",
        },
    )
    assert r.status_code == 200, r.text

    reqs = await client.get(
        f"/cases/{case_id}/system-cam-edit-requests", headers=hdrs
    )
    assert reqs.status_code == 200
    assert reqs.json() == []


async def test_resolve_systemcam_edit_creates_pending_request(client, db, storage):
    _, tok = await _token(db, "disc6@pfl.com", UserRole.AI_ANALYSER)
    hdrs = {"Authorization": f"Bearer {tok}"}
    case_id = await _make_case(client, hdrs)
    await _seed_autocam_extraction(
        db,
        case_id,
        system_cam={"pan": "OLDVAL1234A"},
        cm_cam_il={"pan_number": "NEWVAL1234B"},
    )

    r = await client.post(
        f"/cases/{case_id}/cam-discrepancies/pan/resolve",
        headers=hdrs,
        json={
            "kind": "SYSTEMCAM_EDIT_REQUESTED",
            "comment": "Finpage PAN is stale — branch confirmed the updated PAN.",
            "corrected_value": "NEWVAL1234B",
        },
    )
    assert r.status_code == 200, r.text

    reqs = await client.get(
        f"/cases/{case_id}/system-cam-edit-requests", headers=hdrs
    )
    assert reqs.status_code == 200
    items = reqs.json()
    assert len(items) == 1
    assert items[0]["status"] == "PENDING"
    assert items[0]["requested_system_cam_value"] == "NEWVAL1234B"


async def test_ai_analyser_cannot_decide_edit_request(client, db, storage):
    _, assessor_tok = await _token(db, "disc7a@pfl.com", UserRole.AI_ANALYSER)
    hdrs = {"Authorization": f"Bearer {assessor_tok}"}
    case_id = await _make_case(client, hdrs)
    await _seed_autocam_extraction(
        db,
        case_id,
        system_cam={"pan": "OLDVAL1234A"},
        cm_cam_il={"pan_number": "NEWVAL1234B"},
    )
    r = await client.post(
        f"/cases/{case_id}/cam-discrepancies/pan/resolve",
        headers=hdrs,
        json={
            "kind": "SYSTEMCAM_EDIT_REQUESTED",
            "comment": "Finpage PAN stale — branch confirmed updated PAN.",
            "corrected_value": "NEWVAL1234B",
        },
    )
    assert r.status_code == 200
    reqs = await client.get(
        f"/cases/{case_id}/system-cam-edit-requests", headers=hdrs
    )
    req_id = reqs.json()[0]["id"]

    # Same assessor tries to self-approve → 403
    r2 = await client.post(
        f"/cases/{case_id}/system-cam-edit-requests/{req_id}/decide",
        headers=hdrs,
        json={"approve": True, "decision_comment": "I approve my own request"},
    )
    assert r2.status_code == 403


async def test_admin_can_approve_edit_request(client, db, storage):
    _, assessor_tok = await _token(db, "disc8a@pfl.com", UserRole.AI_ANALYSER)
    _, admin_tok = await _token(db, "disc8b@pfl.com", UserRole.ADMIN)
    hdrs_ass = {"Authorization": f"Bearer {assessor_tok}"}
    hdrs_adm = {"Authorization": f"Bearer {admin_tok}"}
    case_id = await _make_case(client, hdrs_ass)
    await _seed_autocam_extraction(
        db,
        case_id,
        system_cam={"pan": "OLDVAL1234A"},
        cm_cam_il={"pan_number": "NEWVAL1234B"},
    )
    await client.post(
        f"/cases/{case_id}/cam-discrepancies/pan/resolve",
        headers=hdrs_ass,
        json={
            "kind": "SYSTEMCAM_EDIT_REQUESTED",
            "comment": "Finpage PAN stale — please approve update.",
            "corrected_value": "NEWVAL1234B",
        },
    )
    reqs = await client.get(
        f"/cases/{case_id}/system-cam-edit-requests", headers=hdrs_ass
    )
    req_id = reqs.json()[0]["id"]

    decide = await client.post(
        f"/cases/{case_id}/system-cam-edit-requests/{req_id}/decide",
        headers=hdrs_adm,
        json={
            "approve": True,
            "decision_comment": "Confirmed with branch. Approving SystemCam edit.",
        },
    )
    assert decide.status_code == 200, decide.text
    assert decide.json()["status"] == "APPROVED"


async def test_phase1_gate_blocks_on_unresolved_critical(client, db, storage):
    """The core integration test: CRITICAL discrepancy blocks POST /phase1."""
    _, tok = await _token(db, "disc9@pfl.com", UserRole.AI_ANALYSER)
    hdrs = {"Authorization": f"Bearer {tok}"}
    case_id = await _make_case(client, hdrs)
    await _force_stage(db, case_id, CaseStage.INGESTED)
    await _seed_autocam_extraction(
        db,
        case_id,
        system_cam={"pan": "OWLPS6441C"},
        cm_cam_il={"pan_number": "WRONG1234F"},
    )

    r = await client.post(f"/cases/{case_id}/phase1", headers=hdrs)
    assert r.status_code == 409, r.text
    body = r.json()
    # FastAPI wraps the detail dict under "detail"
    detail = body.get("detail", body)
    assert detail.get("reason") == "cam_discrepancies_unresolved"
    pending = detail.get("pending_discrepancies", [])
    assert any(p.get("field_key") == "pan" for p in pending)


async def test_phase1_gate_allows_after_justified_resolution(
    client, db, storage, monkeypatch
):
    """After JUSTIFIED on the sole CRITICAL flag, Phase 1 trigger succeeds.

    We patch the decisioning queue's publish_job to a no-op so the test
    doesn't need a live SQS queue — what we're asserting is only that the
    discrepancy gate is BYPASSED after resolution.
    """
    import app.api.deps as _deps

    async def _noop_publish(payload: dict) -> str:  # noqa: ARG001
        return "test-msg-id"

    class _FakeQueue:
        publish_job = staticmethod(_noop_publish)

    async def _override_dep() -> _FakeQueue:
        return _FakeQueue()

    from app.main import create_app
    app = create_app()
    app.dependency_overrides[_deps.get_decisioning_queue_dep] = _override_dep

    # NB: we reuse the client fixture, so instead of juggling apps we monkey-
    # patch the dep on the already-mounted app the client talks to.
    from app.main import app as live_app  # type: ignore
    live_app.dependency_overrides[_deps.get_decisioning_queue_dep] = _override_dep

    _, tok = await _token(db, "disc10@pfl.com", UserRole.AI_ANALYSER)
    hdrs = {"Authorization": f"Bearer {tok}"}
    case_id = await _make_case(client, hdrs)
    await _force_stage(db, case_id, CaseStage.INGESTED)
    await _seed_autocam_extraction(
        db,
        case_id,
        system_cam={"pan": "OWLPS6441C"},
        cm_cam_il={"pan_number": "WRONG1234F"},
    )
    # Resolve
    resolve = await client.post(
        f"/cases/{case_id}/cam-discrepancies/pan/resolve",
        headers=hdrs,
        json={
            "kind": "JUSTIFIED",
            "comment": "Co-applicant PAN accidentally landed in CM IL; acknowledged.",
        },
    )
    assert resolve.status_code == 200, resolve.text

    # Phase 1 now allowed — gate is passed.
    r = await client.post(f"/cases/{case_id}/phase1", headers=hdrs)
    # cleanup
    live_app.dependency_overrides.pop(_deps.get_decisioning_queue_dep, None)

    assert r.status_code == 202, r.text
    assert "decision_result_id" in r.json()


async def test_report_endpoint_returns_markdown(client, db, storage):
    _, tok = await _token(db, "disc11@pfl.com", UserRole.AI_ANALYSER)
    hdrs = {"Authorization": f"Bearer {tok}"}
    case_id = await _make_case(client, hdrs)
    await _seed_autocam_extraction(
        db,
        case_id,
        system_cam={"foir_overall": 25.35},
        cm_cam_il={"foir": 0.181},
    )
    r = await client.get(
        f"/cases/{case_id}/cam-discrepancies/report", headers=hdrs
    )
    assert r.status_code == 200
    assert "text/markdown" in r.headers["content-type"]
    body = r.text
    assert "CAM Discrepancy Report" in body
    assert "FOIR" in body
    assert "25.35" in body


async def test_report_xlsx_endpoint(client, db, storage):
    import io

    import openpyxl

    _, tok = await _token(db, "disc12@pfl.com", UserRole.AI_ANALYSER)
    hdrs = {"Authorization": f"Bearer {tok}"}
    case_id = await _make_case(client, hdrs)
    await _seed_autocam_extraction(
        db,
        case_id,
        system_cam={"foir_overall": 25.35, "pan": "OWLPS6441C"},
        cm_cam_il={"foir": 0.181, "pan_number": "WRONG1234F"},
    )
    r = await client.get(
        f"/cases/{case_id}/cam-discrepancies/report.xlsx", headers=hdrs
    )
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    # Parse the bytes with openpyxl and sanity-check sheets + values.
    wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True)
    assert set(wb.sheetnames) == {"Summary", "Details"}
    details = wb["Details"]
    rows = list(details.iter_rows(values_only=True))
    header = rows[0]
    assert "Field" in header
    assert "Severity" in header
    body_rows = [r for r in rows[1:] if r[0] is not None]
    assert any("FOIR" in str(r[0]) for r in body_rows)
    assert any("PAN" in str(r[0]) for r in body_rows)
