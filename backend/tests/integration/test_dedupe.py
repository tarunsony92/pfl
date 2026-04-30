"""Integration tests for dedupe service.

Spec §4.5. Tests exact + fuzzy matching, normalization, snapshot parsing.
"""

from __future__ import annotations

from pathlib import Path
from tempfile import TemporaryDirectory

import pytest

from app.enums import DedupeMatchType
from app.worker.dedupe import DedupeResult, run_dedupe
from tests.fixtures.builders.dedupe_builder import build_dedupe_xlsx

# ============================================================================
# Test 1: No snapshot → empty matches + warning
# ============================================================================


@pytest.mark.asyncio
async def test_no_snapshot_returns_empty_with_warning():
    """When snapshot_xlsx_bytes is None, return empty matches with warning."""
    result = await run_dedupe(
        applicant={"aadhaar": "1234567890"},
        co_applicant=None,
        snapshot_xlsx_bytes=None,
    )

    assert isinstance(result, DedupeResult)
    assert result.matches == []
    assert "no_active_snapshot" in result.warnings


# ============================================================================
# Test 2: Exact Aadhaar match
# ============================================================================


@pytest.mark.asyncio
async def test_exact_aadhaar_match():
    """Applicant with matching Aadhaar in snapshot → 1 match, type=AADHAAR, score=1.0."""
    with TemporaryDirectory() as tmpdir:
        customers = [
            {
                "Customer Name": "SEEMA DEVI",
                "Aadhaar": "1234567890",
                "PAN": "ABCDE1234F",
                "Mobile": "9876543210",
                "DOB": "15/03/1985",
                "Address": "123 MG Road, Delhi",
            },
        ]
        xlsx_path = build_dedupe_xlsx(Path(tmpdir) / "dedupe.xlsx", customers)
        xlsx_bytes = xlsx_path.read_bytes()

        result = await run_dedupe(
            applicant={
                "aadhaar": "1234567890",
                "pan": None,
                "mobile": None,
                "name": None,
                "dob": None,
            },
            co_applicant=None,
            snapshot_xlsx_bytes=xlsx_bytes,
        )

        assert len(result.matches) == 1
        match = result.matches[0]
        assert match.match_type == DedupeMatchType.AADHAAR
        assert match.match_score == 1.0
        assert match.matched_customer_id is None  # builder doesn't add customer id


# ============================================================================
# Test 3: Exact PAN match
# ============================================================================


@pytest.mark.asyncio
async def test_exact_pan_match():
    """Applicant with matching PAN in snapshot → 1 match, type=PAN, score=1.0."""
    with TemporaryDirectory() as tmpdir:
        customers = [
            {
                "Customer Name": "RAMESH KUMAR",
                "Aadhaar": "5678901234",
                "PAN": "ABCDE1234F",
                "Mobile": "9123456789",
                "DOB": "22/07/1978",
                "Address": "45 Lajpat Nagar, Delhi",
            },
        ]
        xlsx_path = build_dedupe_xlsx(Path(tmpdir) / "dedupe.xlsx", customers)
        xlsx_bytes = xlsx_path.read_bytes()

        result = await run_dedupe(
            applicant={
                "aadhaar": None,
                "pan": "ABCDE1234F",
                "mobile": None,
                "name": None,
                "dob": None,
            },
            co_applicant=None,
            snapshot_xlsx_bytes=xlsx_bytes,
        )

        assert len(result.matches) == 1
        match = result.matches[0]
        assert match.match_type == DedupeMatchType.PAN
        assert match.match_score == 1.0


# ============================================================================
# Test 4: Exact Mobile match (score=0.9)
# ============================================================================


@pytest.mark.asyncio
async def test_exact_mobile_match():
    """Applicant with matching Mobile in snapshot → 1 match, type=MOBILE, score=0.9."""
    with TemporaryDirectory() as tmpdir:
        customers = [
            {
                "Customer Name": "PRIYA SHARMA",
                "Aadhaar": "9999999999",
                "PAN": "FGHIJ5678K",
                "Mobile": "9876543210",
                "DOB": "10/05/1990",
                "Address": "789 Park Street, Delhi",
            },
        ]
        xlsx_path = build_dedupe_xlsx(Path(tmpdir) / "dedupe.xlsx", customers)
        xlsx_bytes = xlsx_path.read_bytes()

        result = await run_dedupe(
            applicant={
                "aadhaar": None,
                "pan": None,
                "mobile": "9876543210",
                "name": None,
                "dob": None,
            },
            co_applicant=None,
            snapshot_xlsx_bytes=xlsx_bytes,
        )

        assert len(result.matches) == 1
        match = result.matches[0]
        assert match.match_type == DedupeMatchType.MOBILE
        assert match.match_score == 0.9


# ============================================================================
# Test 5: Fuzzy name+DOB match above threshold
# ============================================================================


@pytest.mark.asyncio
async def test_fuzzy_name_dob_match_above_threshold():
    """Applicant with typo in name but matching DOB → fuzzy match with high score."""
    with TemporaryDirectory() as tmpdir:
        customers = [
            {
                "Customer Name": "SEEMA DEVI",
                "Aadhaar": "1234567890",
                "PAN": "ABCDE1234F",
                "Mobile": "9876543210",
                "DOB": "15/03/1985",
                "Address": "123 MG Road",
            },
        ]
        xlsx_path = build_dedupe_xlsx(Path(tmpdir) / "dedupe.xlsx", customers)
        xlsx_bytes = xlsx_path.read_bytes()

        # Applicant has typo: "SEEMA DEV" instead of "SEEMA DEVI"
        result = await run_dedupe(
            applicant={
                "aadhaar": None,
                "pan": None,
                "mobile": None,
                "name": "SEEMA DEV",
                "dob": "15/03/1985",
            },
            co_applicant=None,
            snapshot_xlsx_bytes=xlsx_bytes,
            fuzzy_name_threshold=0.85,
        )

        assert len(result.matches) == 1
        match = result.matches[0]
        assert match.match_type == DedupeMatchType.DOB_NAME
        assert match.match_score >= 0.85


# ============================================================================
# Test 6: Fuzzy name below threshold → no match
# ============================================================================


@pytest.mark.asyncio
async def test_fuzzy_name_below_threshold_no_match():
    """Applicant with completely different name → no fuzzy match."""
    with TemporaryDirectory() as tmpdir:
        customers = [
            {
                "Customer Name": "SEEMA DEVI",
                "Aadhaar": "1234567890",
                "PAN": "ABCDE1234F",
                "Mobile": "9876543210",
                "DOB": "15/03/1985",
                "Address": "123 MG Road",
            },
        ]
        xlsx_path = build_dedupe_xlsx(Path(tmpdir) / "dedupe.xlsx", customers)
        xlsx_bytes = xlsx_path.read_bytes()

        result = await run_dedupe(
            applicant={
                "aadhaar": None,
                "pan": None,
                "mobile": None,
                "name": "RAJESH PANDEY",
                "dob": "15/03/1985",
            },
            co_applicant=None,
            snapshot_xlsx_bytes=xlsx_bytes,
            fuzzy_name_threshold=0.85,
        )

        assert len(result.matches) == 0


# ============================================================================
# Test 7: No matches → empty list
# ============================================================================


@pytest.mark.asyncio
async def test_no_matches_returns_empty_matches():
    """Applicant with no matches in snapshot → empty matches list."""
    with TemporaryDirectory() as tmpdir:
        customers = [
            {
                "Customer Name": "SEEMA DEVI",
                "Aadhaar": "1234567890",
                "PAN": "ABCDE1234F",
                "Mobile": "9876543210",
                "DOB": "15/03/1985",
                "Address": "123 MG Road",
            },
        ]
        xlsx_path = build_dedupe_xlsx(Path(tmpdir) / "dedupe.xlsx", customers)
        xlsx_bytes = xlsx_path.read_bytes()

        result = await run_dedupe(
            applicant={
                "aadhaar": "9999999999",  # different
                "pan": "ZZZZZZZZZZ",  # different
                "mobile": "1111111111",  # different
                "name": "UNKNOWN NAME",  # different
                "dob": "01/01/2000",  # different
            },
            co_applicant=None,
            snapshot_xlsx_bytes=xlsx_bytes,
        )

        assert result.matches == []


# ============================================================================
# Test 8: Co-applicant matching works
# ============================================================================


@pytest.mark.asyncio
async def test_co_applicant_also_checked():
    """Co-applicant is also checked for matches."""
    with TemporaryDirectory() as tmpdir:
        customers = [
            {
                "Customer Name": "JOHN DOE",
                "Aadhaar": "5555555555",
                "PAN": "JJJJJ5555K",
                "Mobile": "9111111111",
                "DOB": "01/01/1980",
                "Address": "999 Some Street",
            },
        ]
        xlsx_path = build_dedupe_xlsx(Path(tmpdir) / "dedupe.xlsx", customers)
        xlsx_bytes = xlsx_path.read_bytes()

        result = await run_dedupe(
            applicant=None,
            co_applicant={
                "aadhaar": "5555555555",
                "pan": None,
                "mobile": None,
                "name": None,
                "dob": None,
            },
            snapshot_xlsx_bytes=xlsx_bytes,
        )

        assert len(result.matches) == 1
        match = result.matches[0]
        assert match.match_type == DedupeMatchType.AADHAAR
        assert match.matched_details["source"] == "co_applicant"


# ============================================================================
# Test 9: Both applicant and co-applicant matches
# ============================================================================


@pytest.mark.asyncio
async def test_both_applicant_and_co_applicant_matches():
    """Both applicant and co-applicant can match."""
    with TemporaryDirectory() as tmpdir:
        customers = [
            {
                "Customer Name": "ALICE SMITH",
                "Aadhaar": "1111111111",
                "PAN": "AAAAA1111B",
                "Mobile": "9000000001",
                "DOB": "11/11/1980",
                "Address": "111 Alice St",
            },
            {
                "Customer Name": "BOB SMITH",
                "Aadhaar": "2222222222",
                "PAN": "BBBBB2222B",
                "Mobile": "9000000002",
                "DOB": "22/22/1981",
                "Address": "222 Bob St",
            },
        ]
        xlsx_path = build_dedupe_xlsx(Path(tmpdir) / "dedupe.xlsx", customers)
        xlsx_bytes = xlsx_path.read_bytes()

        result = await run_dedupe(
            applicant={
                "aadhaar": "1111111111",
                "pan": None,
                "mobile": None,
                "name": None,
                "dob": None,
            },
            co_applicant={
                "aadhaar": "2222222222",
                "pan": None,
                "mobile": None,
                "name": None,
                "dob": None,
            },
            snapshot_xlsx_bytes=xlsx_bytes,
        )

        assert len(result.matches) == 2
        sources = [m.matched_details["source"] for m in result.matches]
        assert "applicant" in sources
        assert "co_applicant" in sources


# ============================================================================
# Test 10: Empty fields in applicant don't match empty fields in snapshot
# ============================================================================


@pytest.mark.asyncio
async def test_empty_input_fields_dont_match_empty_snapshot_fields():
    """Blank Aadhaar in applicant should not match blank Aadhaar in snapshot."""
    with TemporaryDirectory() as tmpdir:
        customers = [
            {
                "Customer Name": "NO AADHAAR PERSON",
                "Aadhaar": "",  # empty
                "PAN": "NOAA0000AA",
                "Mobile": "9888888888",
                "DOB": "03/03/1975",
                "Address": "No Aadhaar Road",
            },
        ]
        xlsx_path = build_dedupe_xlsx(Path(tmpdir) / "dedupe.xlsx", customers)
        xlsx_bytes = xlsx_path.read_bytes()

        result = await run_dedupe(
            applicant={"aadhaar": "", "pan": None, "mobile": None, "name": None, "dob": None},
            co_applicant=None,
            snapshot_xlsx_bytes=xlsx_bytes,
        )

        # Should not match (both empty)
        assert len(result.matches) == 0


# ============================================================================
# Test 11: Normalization handles spaces and case
# ============================================================================


@pytest.mark.asyncio
async def test_normalization_handles_spaces_and_case():
    """PAN with spaces and lowercase should match uppercase PAN."""
    with TemporaryDirectory() as tmpdir:
        customers = [
            {
                "Customer Name": "NORM TEST",
                "Aadhaar": "7777777777",
                "PAN": "ABCDE1234F",  # uppercase, no spaces
                "Mobile": "9777777777",
                "DOB": "07/07/1970",
                "Address": "Normalization St",
            },
        ]
        xlsx_path = build_dedupe_xlsx(Path(tmpdir) / "dedupe.xlsx", customers)
        xlsx_bytes = xlsx_path.read_bytes()

        # Input PAN: lowercase + spaces
        result = await run_dedupe(
            applicant={
                "aadhaar": None,
                "pan": "  abcde 1234 f  ",
                "mobile": None,
                "name": None,
                "dob": None,
            },
            co_applicant=None,
            snapshot_xlsx_bytes=xlsx_bytes,
        )

        assert len(result.matches) == 1
        match = result.matches[0]
        assert match.match_type == DedupeMatchType.PAN


# ============================================================================
# Test 12: Aadhaar with spaces matches plain Aadhaar
# ============================================================================


@pytest.mark.asyncio
async def test_aadhaar_with_spaces_matches_plain():
    """Aadhaar with spaces (1234 5678 9012) should match plain (123456789012)."""
    with TemporaryDirectory() as tmpdir:
        customers = [
            {
                "Customer Name": "SPACE TEST",
                "Aadhaar": "123456789012",  # plain
                "PAN": "SPAA0000ST",
                "Mobile": "9600000000",
                "DOB": "06/06/1965",
                "Address": "Space Ave",
            },
        ]
        xlsx_path = build_dedupe_xlsx(Path(tmpdir) / "dedupe.xlsx", customers)
        xlsx_bytes = xlsx_path.read_bytes()

        # Input with spaces
        result = await run_dedupe(
            applicant={
                "aadhaar": "1234 5678 9012",
                "pan": None,
                "mobile": None,
                "name": None,
                "dob": None,
            },
            co_applicant=None,
            snapshot_xlsx_bytes=xlsx_bytes,
        )

        assert len(result.matches) == 1
        match = result.matches[0]
        assert match.match_type == DedupeMatchType.AADHAAR
        assert match.match_score == 1.0


# ============================================================================
# Test 13: Multiple field matches on same customer (e.g., both Aadhaar and PAN)
# ============================================================================


@pytest.mark.asyncio
async def test_multiple_fields_match_same_customer():
    """Both Aadhaar and PAN match the same customer → 2 matches returned."""
    with TemporaryDirectory() as tmpdir:
        customers = [
            {
                "Customer Name": "MULTI MATCH",
                "Aadhaar": "4444444444",
                "PAN": "MMMM0000MM",
                "Mobile": "9444444444",
                "DOB": "04/04/1984",
                "Address": "Multi St",
            },
        ]
        xlsx_path = build_dedupe_xlsx(Path(tmpdir) / "dedupe.xlsx", customers)
        xlsx_bytes = xlsx_path.read_bytes()

        result = await run_dedupe(
            applicant={
                "aadhaar": "4444444444",
                "pan": "MMMM0000MM",
                "mobile": None,
                "name": None,
                "dob": None,
            },
            co_applicant=None,
            snapshot_xlsx_bytes=xlsx_bytes,
        )

        # Both Aadhaar and PAN should match
        assert len(result.matches) == 2
        types = {m.match_type for m in result.matches}
        assert DedupeMatchType.AADHAAR in types
        assert DedupeMatchType.PAN in types


# ============================================================================
# Test 14: Empty snapshot (no data rows) returns warning
# ============================================================================


@pytest.mark.asyncio
async def test_empty_snapshot_data_rows_returns_warning():
    """Snapshot with header but no data rows → warning and empty matches."""
    from tempfile import TemporaryDirectory

    import openpyxl

    with TemporaryDirectory():
        # Create xlsx with only header, no data rows
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Customer_Dedupe"
        ws["A1"] = "Customer Name"
        ws["B1"] = "Aadhaar"
        ws["C1"] = "PAN"
        ws["D1"] = "Mobile"
        ws["E1"] = "DOB"
        ws["F1"] = "Address"

        xlsx_path = "/tmp/empty_dedupe.xlsx"
        wb.save(xlsx_path)

        with open(xlsx_path, "rb") as f:
            xlsx_bytes = f.read()

        result = await run_dedupe(
            applicant={
                "aadhaar": "1234567890",
                "pan": None,
                "mobile": None,
                "name": None,
                "dob": None,
            },
            co_applicant=None,
            snapshot_xlsx_bytes=xlsx_bytes,
        )

        assert result.matches == []
        assert "snapshot_has_no_data_rows" in result.warnings


# ============================================================================
# Test 15: Invalid xlsx file returns warning
# ============================================================================


@pytest.mark.asyncio
async def test_invalid_xlsx_bytes_returns_warning():
    """Passing invalid xlsx bytes → warning and empty matches."""
    result = await run_dedupe(
        applicant={"aadhaar": "1234567890", "pan": None, "mobile": None, "name": None, "dob": None},
        co_applicant=None,
        snapshot_xlsx_bytes=b"not a valid xlsx file",
    )

    assert result.matches == []
    assert any("failed_to_load_xlsx" in w for w in result.warnings)


# ============================================================================
# Test 16: Both subjects None still returns result
# ============================================================================


@pytest.mark.asyncio
async def test_both_subjects_none_returns_empty():
    """Both applicant and co_applicant are None → empty matches (no warning)."""
    with TemporaryDirectory() as tmpdir:
        customers = [
            {
                "Customer Name": "SOMEONE",
                "Aadhaar": "1111111111",
                "PAN": "AAAA0000AA",
                "Mobile": "9111111111",
                "DOB": "01/01/1990",
                "Address": "123 St",
            },
        ]
        xlsx_path = build_dedupe_xlsx(Path(tmpdir) / "dedupe.xlsx", customers)
        xlsx_bytes = xlsx_path.read_bytes()

        result = await run_dedupe(
            applicant=None,
            co_applicant=None,
            snapshot_xlsx_bytes=xlsx_bytes,
        )

        assert result.matches == []
        # No data_rows warning since subjects are None
        assert "snapshot_has_no_data_rows" not in result.warnings


# ============================================================================
# Test 17: Snapshot with empty header row
# ============================================================================


@pytest.mark.asyncio
async def test_empty_header_row_returns_warning():
    """Snapshot with empty header row → warning and empty matches."""
    from tempfile import TemporaryDirectory

    import openpyxl

    with TemporaryDirectory():
        # Create xlsx with only blank header
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Customer_Dedupe"
        # No header values - all None

        xlsx_path = "/tmp/blank_header_dedupe.xlsx"
        wb.save(xlsx_path)

        with open(xlsx_path, "rb") as f:
            xlsx_bytes = f.read()

        result = await run_dedupe(
            applicant={
                "aadhaar": "1234567890",
                "pan": None,
                "mobile": None,
                "name": None,
                "dob": None,
            },
            co_applicant=None,
            snapshot_xlsx_bytes=xlsx_bytes,
        )

        assert result.matches == []
        assert "empty_header_row" in result.warnings


# ============================================================================
# Test 18: Non-matching sheet name falls back to active sheet
# ============================================================================


@pytest.mark.asyncio
async def test_non_customer_dedupe_sheet_falls_back_to_active():
    """Sheet with non-matching name falls back to active sheet."""
    from tempfile import TemporaryDirectory

    import openpyxl

    with TemporaryDirectory():
        # Create xlsx with different sheet name (not Customer_Dedupe)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Other_Sheet"  # not "Customer_Dedupe"
        ws["A1"] = "Customer Name"
        ws["B1"] = "Aadhaar"
        ws["C1"] = "PAN"
        ws["D1"] = "Mobile"
        ws["E1"] = "DOB"
        ws["F1"] = "Address"
        ws["A2"] = "TEST PERSON"
        ws["B2"] = "1234567890"
        ws["C2"] = "ABCD0000AB"
        ws["D2"] = "9000000000"
        ws["E2"] = "01/01/1990"
        ws["F2"] = "Test Address"

        xlsx_path = "/tmp/other_sheet_dedupe.xlsx"
        wb.save(xlsx_path)

        with open(xlsx_path, "rb") as f:
            xlsx_bytes = f.read()

        result = await run_dedupe(
            applicant={
                "aadhaar": "1234567890",
                "pan": None,
                "mobile": None,
                "name": None,
                "dob": None,
            },
            co_applicant=None,
            snapshot_xlsx_bytes=xlsx_bytes,
        )

        # Should match because it falls back to active sheet
        assert len(result.matches) == 1
        assert result.matches[0].match_type == DedupeMatchType.AADHAAR


# ============================================================================
# Test 19: Header with None/mixed values handles correctly
# ============================================================================


@pytest.mark.asyncio
async def test_header_with_none_values_skips_none():
    """Header row with mixed None and actual values parses correctly."""
    from tempfile import TemporaryDirectory

    import openpyxl

    with TemporaryDirectory():
        # Create xlsx with None values mixed into header
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Customer_Dedupe"
        # Put headers with some None mixed in
        ws["A1"] = "Customer Name"
        ws["B1"] = None  # None value
        ws["C1"] = "Aadhaar"
        ws["D1"] = "PAN"
        ws["E1"] = None  # Another None
        ws["F1"] = "Mobile"
        ws["G1"] = "DOB"
        ws["H1"] = "Address"

        ws["A2"] = "TEST PERSON"
        ws["C2"] = "1234567890"
        ws["D2"] = "ABCD0000AB"
        ws["F2"] = "9000000000"
        ws["G2"] = "01/01/1990"
        ws["H2"] = "Test Address"

        xlsx_path = "/tmp/none_header_dedupe.xlsx"
        wb.save(xlsx_path)

        with open(xlsx_path, "rb") as f:
            xlsx_bytes = f.read()

        result = await run_dedupe(
            applicant={
                "aadhaar": "1234567890",
                "pan": None,
                "mobile": None,
                "name": None,
                "dob": None,
            },
            co_applicant=None,
            snapshot_xlsx_bytes=xlsx_bytes,
        )

        # Should match Aadhaar even with None values in header
        assert len(result.matches) == 1
        assert result.matches[0].match_type == DedupeMatchType.AADHAAR
