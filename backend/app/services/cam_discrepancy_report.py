"""XLSX exporter for CAM discrepancy reports.

The markdown exporter lives in app/services/cam_discrepancy.py (keeps the
report text embedded with the rest of the discrepancy service for easy
reuse by tests + future docx generation). This module is the xlsx-specific
layer: takes the same CamDiscrepancySummary and writes a two-sheet
workbook suitable for credit-ops review.

Sheet 1 — "Summary":
  Case ID | Loan ID | Generated At | Total Fields | Unresolved Critical |
  Unresolved Warning | Phase 1 Blocked?

Sheet 2 — "Details":
  Field | Severity | SystemCam (finpage) | CM CAM IL (manual) |
  Diff (abs) | Diff (%) | Tolerance | Resolution Kind |
  Corrected Value | Assessor Comment | Resolved By | Resolved At |
  Edit Request Status | Edit Request Value | Edit Request Justification
"""

from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.schemas.cam_discrepancy import CamDiscrepancySummary


_HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_CRITICAL_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
_WARNING_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
_RESOLVED_FILL = PatternFill(start_color="D1FAE5", end_color="D1FAE5", fill_type="solid")


def build_xlsx(summary: CamDiscrepancySummary, case_loan_id: str | None = None) -> bytes:
    wb = Workbook()

    # --- Summary sheet ---
    ws_sum = wb.active
    ws_sum.title = "Summary"
    sum_headers = [
        "Case ID",
        "Loan ID",
        "Generated At",
        "Total Fields",
        "Unresolved Critical",
        "Unresolved Warning",
        "Phase 1 Blocked",
    ]
    for col, h in enumerate(sum_headers, start=1):
        c = ws_sum.cell(row=1, column=col, value=h)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = Alignment(horizontal="left")

    ws_sum.cell(row=2, column=1, value=str(summary.case_id))
    ws_sum.cell(row=2, column=2, value=case_loan_id or "—")
    ws_sum.cell(row=2, column=3, value=summary.generated_at.isoformat())
    ws_sum.cell(row=2, column=4, value=summary.total)
    ws_sum.cell(row=2, column=5, value=summary.unresolved_critical)
    ws_sum.cell(row=2, column=6, value=summary.unresolved_warning)
    ws_sum.cell(row=2, column=7, value="YES" if summary.phase1_blocked else "no")

    for col in range(1, len(sum_headers) + 1):
        ws_sum.column_dimensions[get_column_letter(col)].width = 22

    # --- Details sheet ---
    ws = wb.create_sheet("Details")
    detail_headers = [
        "Field",
        "Severity",
        "SystemCam (finpage)",
        "CM CAM IL (manual)",
        "Diff (abs)",
        "Diff (%)",
        "Tolerance",
        "Why flagged",
        "Resolution Kind",
        "Corrected Value",
        "Assessor Comment",
        "Resolved By",
        "Resolved At",
        "Edit Request Status",
        "Requested Value",
        "Edit Justification",
    ]
    for col, h in enumerate(detail_headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.fill = _HEADER_FILL
        c.font = _HEADER_FONT
        c.alignment = Alignment(horizontal="left")

    for idx, view in enumerate(summary.views, start=2):
        flag = view.flag
        res = view.resolution
        er = view.pending_edit_request

        row_values = [
            view.field_label,
            flag.severity.value if flag else "—",
            flag.system_cam_value if flag else (res.system_cam_value_at_resolve if res else None),
            flag.cm_cam_il_value if flag else (res.cm_cam_il_value_at_resolve if res else None),
            flag.diff_abs if flag else None,
            flag.diff_pct if flag else None,
            flag.tolerance_description if flag else "",
            flag.note if flag else "",
            res.kind.value if res else "",
            res.corrected_value if res else "",
            res.comment if res else "",
            str(res.resolved_by) if res else "",
            res.resolved_at.isoformat() if res else "",
            er.status.value if er else "",
            er.requested_system_cam_value if er else "",
            er.justification if er else "",
        ]
        for col_idx, val in enumerate(row_values, start=1):
            c = ws.cell(row=idx, column=col_idx, value=val)
            c.alignment = Alignment(vertical="top", wrap_text=True)

        # Row background — severity + resolution state
        if res is not None:
            fill = _RESOLVED_FILL
        elif flag and flag.severity.value == "CRITICAL":
            fill = _CRITICAL_FILL
        elif flag and flag.severity.value == "WARNING":
            fill = _WARNING_FILL
        else:
            fill = None
        if fill is not None:
            for col in range(1, len(detail_headers) + 1):
                ws.cell(row=idx, column=col).fill = fill

    # Column widths
    widths = [24, 10, 22, 22, 10, 10, 28, 36, 22, 22, 48, 36, 22, 18, 22, 48]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
