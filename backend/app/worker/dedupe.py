"""Dedupe service: exact + fuzzy matching against Customer_Dedupe snapshot.

Spec §4.5. No DB calls — pure in-memory matching logic.
"""

from __future__ import annotations

import re
from dataclasses import dataclass, field
from io import BytesIO
from typing import Any

import openpyxl
from rapidfuzz.fuzz import token_sort_ratio

from app.enums import DedupeMatchType


@dataclass
class DedupeMatchCandidate:
    match_type: DedupeMatchType
    match_score: float
    matched_customer_id: str | None
    matched_details: dict[str, Any]


@dataclass
class DedupeResult:
    matches: list[DedupeMatchCandidate] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


def _normalize_aadhaar(v: Any) -> str:
    if not v or (isinstance(v, str) and not v.strip()):
        return ""
    return re.sub(r"\D", "", str(v).strip())


def _normalize_pan(v: Any) -> str:
    if not v or (isinstance(v, str) and not v.strip()):
        return ""
    return re.sub(r"[^A-Z0-9]", "", str(v).strip().upper())


def _normalize_mobile(v: Any) -> str:
    if not v or (isinstance(v, str) and not v.strip()):
        return ""
    digits = re.sub(r"\D", "", str(v).strip())
    return digits[-10:] if len(digits) >= 10 else ""


def _normalize_name(v: Any) -> str:
    if not v or (isinstance(v, str) and not v.strip()):
        return ""
    s = str(v).strip().upper()
    s = re.sub(r"\s+", " ", s)
    return re.sub(r"[^A-Z0-9 ]", "", s).strip()


def _find_col(header: list[Any], keywords: list[str]) -> int | None:
    for i, h in enumerate(header):
        if h and any(kw in str(h).strip().lower() for kw in keywords):
            return i
    return None


# Declared async for consistency with the worker pipeline that awaits it;
# body is pure in-memory CPU (openpyxl + rapidfuzz) with no IO.
async def run_dedupe(
    *,
    applicant: dict[str, Any] | None,
    co_applicant: dict[str, Any] | None,
    snapshot_xlsx_bytes: bytes | None,
    fuzzy_name_threshold: float = 0.85,
) -> DedupeResult:
    result = DedupeResult()

    if not snapshot_xlsx_bytes:
        result.warnings.append("no_active_snapshot")
        return result

    try:
        wb = openpyxl.load_workbook(BytesIO(snapshot_xlsx_bytes), data_only=True)
    except Exception as e:
        result.warnings.append(f"failed_to_load_xlsx: {e}")
        return result

    ws = None
    for sheet in wb.sheetnames:
        if sheet.lower() == "customer_dedupe":
            ws = wb[sheet]
            break
    ws = ws or wb.active

    header = [c.value for c in ws[1]]
    if not header or all(h is None for h in header):
        result.warnings.append("empty_header_row")
        return result

    cols = {
        "id": _find_col(header, ["customer id", "cust id"]),
        "name": _find_col(header, ["customer name", "name"]),
        "aadhaar": _find_col(header, ["aadhaar", "aadhaar id"]),
        "pan": _find_col(header, ["pan", "pan card"]),
        "mobile": _find_col(header, ["mobile", "mobile no", "phone"]),
        "dob": _find_col(header, ["dob", "date of birth"]),
        "address": _find_col(header, ["address"]),
    }

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(
            {k: row[v] if v is not None and v < len(row) else None for k, v in cols.items()}
        )

    if not rows:
        result.warnings.append("snapshot_has_no_data_rows")
        return result

    for subj, stype in [(applicant, "applicant"), (co_applicant, "co_applicant")]:
        if not subj:
            continue

        saa = _normalize_aadhaar(subj.get("aadhaar"))
        spa = _normalize_pan(subj.get("pan"))
        smo = _normalize_mobile(subj.get("mobile"))
        sna = _normalize_name(subj.get("name"))
        sdo = str(subj.get("dob")).strip() if subj.get("dob") else ""

        for row in rows:
            rid = row.get("id")
            raa = _normalize_aadhaar(row.get("aadhaar"))
            rpa = _normalize_pan(row.get("pan"))
            rmo = _normalize_mobile(row.get("mobile"))
            rna = _normalize_name(row.get("name"))
            rdo = str(row.get("dob")).strip() if row.get("dob") else ""

            details = {
                "source": stype,
                "customer_id": rid,
                "customer_name": row.get("name"),
                "aadhaar": row.get("aadhaar"),
                "pan": row.get("pan"),
                "mobile": row.get("mobile"),
                "dob": row.get("dob"),
                "address": row.get("address"),
            }

            if saa and raa and saa == raa:
                result.matches.append(
                    DedupeMatchCandidate(
                        DedupeMatchType.AADHAAR, 1.0, str(rid) if rid else None, details
                    )
                )

            if spa and rpa and spa == rpa:
                result.matches.append(
                    DedupeMatchCandidate(
                        DedupeMatchType.PAN, 1.0, str(rid) if rid else None, details
                    )
                )

            if smo and rmo and smo == rmo:
                # Mobile is high-confidence but not definitive (spec §4.5) → 0.9
                result.matches.append(
                    DedupeMatchCandidate(
                        DedupeMatchType.MOBILE, 0.9, str(rid) if rid else None, details
                    )
                )

            if sna and rna and sdo and rdo:
                score = token_sort_ratio(f"{sna} {sdo}", f"{rna} {rdo}") / 100.0
                if score >= fuzzy_name_threshold:
                    result.matches.append(
                        DedupeMatchCandidate(
                            DedupeMatchType.DOB_NAME, score, str(rid) if rid else None, details
                        )
                    )

    return result
