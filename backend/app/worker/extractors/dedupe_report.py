"""Extractor for DEDUPE_REPORT artefact (Finpage Customer_Dedupe xlsx).

Two known shapes:
  1. Empty (header-only, no data rows) -> row_count = 0, identity unique on file.
  2. Non-empty (header + data rows) -> row_count > 0, surfaces matched
     identity records to the L5.5 orchestrator for assessor review.

Header row may be preceded by a title row like "Finpage - Home"; the
extractor scans the first 5 rows for one whose cells include any string
starting with "customer" (case-insensitive) AND has at least 3 non-empty
cells (to exclude single-cell metadata rows like "Customer reviewed at").
"""

from __future__ import annotations

import io
import re
from typing import Any

import openpyxl

from app.enums import ExtractionStatus
from app.worker.extractors.base import BaseExtractor, ExtractionResult


class DedupeReportExtractor(BaseExtractor):
    extractor_name = "dedupe_report"
    schema_version = "1.0"

    _NON_ALNUM = re.compile(r"[^a-z0-9]+")

    async def extract(self, filename: str, body_bytes: bytes) -> ExtractionResult:
        try:
            wb = openpyxl.load_workbook(
                io.BytesIO(body_bytes), read_only=True, data_only=True
            )
        except Exception as exc:  # noqa: BLE001
            return ExtractionResult(
                status=ExtractionStatus.FAILED,
                schema_version=self.schema_version,
                data={},
                error_message=f"openpyxl load failed: {exc}",
            )

        ws = wb.worksheets[0]
        header_row_idx, headers = self._locate_header(ws)
        if header_row_idx is None:
            return ExtractionResult(
                status=ExtractionStatus.FAILED,
                schema_version=self.schema_version,
                data={},
                error_message="Header row not found (no row contained a 'Customer ...' cell in the first 5 rows)",
            )

        rows: list[dict[str, Any]] = []
        for raw in ws.iter_rows(
            min_row=header_row_idx + 1, values_only=True
        ):
            cells = list(raw)
            # Skip rows where ALL cells are blank (trailing empty rows).
            if all(c in (None, "") for c in cells):
                continue
            normalised: dict[str, Any] = {}
            for i, h in enumerate(headers):
                if not h:
                    continue
                key = self._normalise_key(h)
                value = cells[i] if i < len(cells) else None
                normalised[key] = value
            rows.append(normalised)

        matched_fields = sorted({
            k for r in rows for k, v in r.items() if v not in (None, "")
        })

        return ExtractionResult(
            status=ExtractionStatus.SUCCESS,
            schema_version=self.schema_version,
            data={
                "row_count": len(rows),
                "matched_rows": rows,
                "matched_fields": matched_fields,
            },
        )

    @staticmethod
    def _locate_header(ws: Any) -> tuple[int | None, list[str]]:
        """Scan first 5 rows for a tabular header.

        Heuristic: a real header row has at least one cell starting with
        "customer" (case-insensitive) AND at least 3 non-empty cells. The
        second clause excludes single-cell metadata rows like
        "Customer reviewed at" or "Customer cases — generated at ...".
        """
        for row_idx in range(1, 6):
            try:
                cells_raw = next(
                    ws.iter_rows(min_row=row_idx, max_row=row_idx, values_only=True)
                )
            except StopIteration:
                break
            cells = [str(c).strip() if c is not None else "" for c in cells_raw]
            non_empty = sum(1 for c in cells if c)
            has_customer = any(c.lower().startswith("customer") for c in cells)
            if has_customer and non_empty >= 3:
                return row_idx, cells
        return None, []

    @staticmethod
    def _normalise_key(header: str) -> str:
        """Snake_case normaliser. 'Customer Id' -> 'customer_id',
        'Aadhaar #' -> 'aadhaar', 'Voter (Card)' -> 'voter_card'."""
        s = DedupeReportExtractor._NON_ALNUM.sub("_", header.strip().lower())
        return s.strip("_")
