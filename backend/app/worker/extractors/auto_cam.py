"""AutoCam extractor — parses the 4-sheet Auto CAM xlsx produced by PFL ops.

Fix (M4): fuzzy sheet-name matching (substring) + fuzzy label matching (substring
on normalised text) so real-world files with varied naming still extract data.
The synthetic test fixture uses exact names and exact labels so it is unaffected.
"""

from __future__ import annotations

import io
from typing import Any

import openpyxl

from app.enums import ExtractionStatus
from app.worker.extractors.base import BaseExtractor, ExtractionResult

# ---------------------------------------------------------------------------
# Sheet configuration
# ---------------------------------------------------------------------------

# Canonical sheet names (used as output keys and for exact matching first).
_CANONICAL_SHEETS = ("SystemCam", "Elegibilty", "CM CAM IL", "Health Sheet")

# Substring aliases: if a workbook sheet name (normalised) contains any of these
# substrings it is mapped to the canonical sheet.
_SHEET_ALIASES: dict[str, list[str]] = {
    "SystemCam": ["systemcam", "system cam", "system_cam", "cam_report"],
    "Elegibilty": ["elegibilty", "eligib", "elig"],
    "CM CAM IL": ["cm cam il", "cmcam", "cm_cam", "ilcam"],
    "Health Sheet": ["health sheet", "health_sheet", "healthsheet"],
}

# Output dict keys corresponding to each canonical sheet
_SHEET_DATA_KEYS = {
    "SystemCam": "system_cam",
    "Elegibilty": "eligibility",
    "CM CAM IL": "cm_cam_il",
    "Health Sheet": "health_sheet",
}

# ---------------------------------------------------------------------------
# Label → field key mappings per sheet
# Each mapping entry is (normalised_substring_or_exact, field_key).
# Matching: exact normalised match first, then substring match in order.
# ---------------------------------------------------------------------------

# Exact labels (normalised)  ─────────────────────────────────────────────────
_SHEET_EXACT_MAPS: dict[str, dict[str, str]] = {
    "SystemCam": {
        "applicant name": "applicant_name",
        "date of birth": "date_of_birth",
        "pan": "pan",
        "loan amount": "loan_amount",
    },
    "Elegibilty": {
        "cibil score": "cibil_score",
        "foir": "foir",
        "eligible amount": "eligible_amount",
    },
    "CM CAM IL": {
        "borrower name": "borrower_name",
        "pan number": "pan_number",
        "loan required": "loan_required",
        "cibil": "cibil",
        "hosehold expanses": "household_expense",  # real CAM misspelling
        "household expense": "household_expense",
        "household expenses": "household_expense",
        "emi obligation": "emi_obligation",
        "disposable income": "disposable_income",
        "servable emi": "servable_emi",
    },
    "Health Sheet": {
        "total monthly income": "total_monthly_income",
        "total monthly expense": "total_monthly_expense",
        "net surplus": "net_surplus",
    },
}

# Fuzzy substring patterns  ──────────────────────────────────────────────────
# Order matters: more-specific patterns before broader ones within the same sheet.
_SHEET_FUZZY_MAPS: dict[str, list[tuple[str, str]]] = {
    "SystemCam": [
        ("first name", "applicant_name"),       # SystemCam uses "First Name" in real file
        ("name of applicant", "applicant_name"),
        ("applicant", "applicant_name"),
        ("date of birth", "date_of_birth"),
        ("dob", "date_of_birth"),
        ("pan", "pan"),
        ("loan amount", "loan_amount"),
        ("expected emi", "expected_emi"),
        ("foir installment", "foir_installment"),
        ("foir overall", "foir_overall"),
        ("own installment", "own_installment"),
        ("other loan instal", "other_loan_installments"),
    ],
    "Elegibilty": [
        ("cibil score", "cibil_score"),
        ("highmark score", "cibil_score"),      # use highmark as fallback if no cibil
        ("foir", "foir"),
        ("eligible amount", "eligible_amount"),
        ("loan amount recommended", "eligible_amount"),
        ("name of applicant", "applicant_name"),
        ("applicant", "applicant_name"),
    ],
    "CM CAM IL": [
        ("borrower name", "borrower_name"),
        ("name of applicant", "borrower_name"),
        ("pan number", "pan_number"),
        ("loan required", "loan_required"),
        ("cibil", "cibil"),
        ("total income", "total_monthly_income"),
        ("hosehold expanses", "household_expense"),
        ("household expanses", "household_expense"),
        ("household expense", "household_expense"),
        ("household expenses", "household_expense"),
        ("emi obligation", "emi_obligation"),
        ("disposable income", "disposable_income"),
        ("servable emi", "servable_emi"),
        ("foir", "foir"),
    ],
    "Health Sheet": [
        ("total monthly income", "total_monthly_income"),
        ("total income", "total_monthly_income"),
        ("total monthly expense", "total_monthly_expense"),
        ("net surplus", "net_surplus"),
        ("foir", "foir"),
    ],
}

# Right-column fuzzy lists — used ONLY when scanning cols F/G/H. Narrow to
# finance-section labels so we don't accidentally match applicant identity
# fields from the Co-Borrower Details block (which lives in col F rows 1-91
# of the CAM_REPORT layout). Finance labels (rows 92+) are unambiguous.
_SHEET_RIGHT_COL_FUZZY: dict[str, list[tuple[str, str]]] = {
    "SystemCam": [
        ("total business expense", "total_business_expense"),
        ("total household expense", "total_household_expense"),
        ("total business income", "total_business_income"),
        ("total household income", "total_household_income"),
        ("household expenses", "household_expense"),
        ("net business income", "net_business_income"),
        ("net household income", "net_household_income"),
        ("net margin", "net_margin"),
        ("dscr", "dscr"),
        ("foir overall", "foir_overall"),
        ("foir installment", "foir_installment"),
    ],
}


# ---------------------------------------------------------------------------
# Sheet name resolution
# ---------------------------------------------------------------------------


def _resolve_sheet(workbook_names: list[str]) -> dict[str, str]:
    """Return a mapping canonical_name → actual_workbook_sheet_name.

    First tries exact match, then substring alias match.
    """
    normalised = {name: " ".join(name.strip().lower().split()) for name in workbook_names}
    mapping: dict[str, str] = {}
    for canonical in _CANONICAL_SHEETS:
        # Exact match first
        if canonical in workbook_names:
            mapping[canonical] = canonical
            continue
        # Alias match
        for actual, norm in normalised.items():
            aliases = _SHEET_ALIASES.get(canonical, [])
            if any(alias in norm for alias in aliases):
                mapping[canonical] = actual
                break
    return mapping


# ---------------------------------------------------------------------------
# Row parsing
# ---------------------------------------------------------------------------


def _parse_sheet(  # noqa: E501
    ws: Any,
    exact_map: dict[str, str],
    fuzzy_list: list[tuple[str, str]],
    *,
    right_fuzzy_list: list[tuple[str, str]] | None = None,
) -> dict[str, Any]:
    """Iterate all rows of *ws*, matching labels against known field keys.

    Checks columns A+B (standard fixture layout) and also columns B+C
    (real CAM layout where col A is a section heading and labels are in B).
    Exact normalised match wins; then fuzzy substring match.
    """
    result: dict[str, Any] = {}

    def _normalise(raw: Any) -> str:
        return " ".join(str(raw).strip().split()).lower() if raw is not None else ""

    def _resolve(normalised_label: str) -> str | None:
        # Exact
        key = exact_map.get(normalised_label)
        if key:
            return key
        # Fuzzy substring
        for pattern, field_key in fuzzy_list:
            if pattern in normalised_label:
                return field_key
        return None

    def _resolve_right(normalised_label: str) -> str | None:
        if not right_fuzzy_list:
            return None
        # Right-column labels are always "new" in the finance section, so we
        # don't bother with exact_map here — the right fuzzy list covers it.
        for pattern, field_key in right_fuzzy_list:
            if pattern in normalised_label:
                return field_key
        return None

    def _is_empty(v: Any) -> bool:
        return v is None or (isinstance(v, str) and not v.strip())

    def _looks_like_header(v: Any) -> bool:
        """True if *v* reads like a header row cell (e.g. 'Co-Applicant Details').

        Used to skip rows where the value column actually holds another header.
        """
        if not isinstance(v, str):
            return False
        norm = " ".join(v.strip().split()).lower()
        return (
            norm.endswith(" details")
            or norm.endswith(" particulars")
            or norm in {"applicant", "co-applicant", "particulars", "details"}
        )

    # Scan up to 200 rows. Widen to col 10 so we can also pick up labels in the
    # right half of the CAM_REPORT single-sheet layout (labels in col F / index 5,
    # values in col H / index 7).
    for row in ws.iter_rows(min_row=1, max_row=200, min_col=1, max_col=10, values_only=True):
        col_a = row[0] if len(row) > 0 else None
        col_b = row[1] if len(row) > 1 else None
        col_c = row[2] if len(row) > 2 else None
        col_f = row[5] if len(row) > 5 else None
        col_g = row[6] if len(row) > 6 else None
        col_h = row[7] if len(row) > 7 else None

        # (A, B) pair — standard fixture layout; also A-label with C fallback
        # when B is empty (real CAM SystemCam: col A = label, col B = empty, col C = value)
        if col_a is not None:
            norm = _normalise(col_a)
            key = _resolve(norm)
            if key and key not in result:
                value = col_b if not _is_empty(col_b) else col_c
                if not _is_empty(value) and not _looks_like_header(value):
                    result[key] = value

        # (B, C) pair — real CAM Elegibilty/CM CAM IL layout where A is a section
        # heading, B holds the label, C holds the applicant value
        if col_b is not None:
            norm = _normalise(col_b)
            key = _resolve(norm)
            if (
                key
                and key not in result
                and not _is_empty(col_c)
                and not _looks_like_header(col_c)
            ):
                result[key] = col_c

        # (F, G/H) pair — right half of CAM_REPORT / SystemCam layout where the
        # right side carries a second label-value pair (e.g. Total Household
        # Expense in col F, value in col H). We only match a narrow finance
        # vocabulary here so we don't pick up Co-Borrower identity labels like
        # "PAN" or "Name" that also sit in col F of rows 1-91.
        if col_f is not None and right_fuzzy_list:
            norm = _normalise(col_f)
            key = _resolve_right(norm)
            if key and key not in result:
                value = col_g if not _is_empty(col_g) else col_h
                if not _is_empty(value) and not _looks_like_header(value):
                    result[key] = value

    return result


# ---------------------------------------------------------------------------
# Cross-sheet lookup helper
# ---------------------------------------------------------------------------


def _find_in_data(data: dict[str, Any], field_keys: tuple[str, ...]) -> Any:
    """Return the first non-empty value matching any of *field_keys* across all sheets."""
    for sheet_data in data.values():
        if not isinstance(sheet_data, dict):
            continue
        for k in field_keys:
            v = sheet_data.get(k)
            if v is None:
                continue
            if isinstance(v, str) and not v.strip():
                continue
            return v
    return None


# ---------------------------------------------------------------------------
# Extractor
# ---------------------------------------------------------------------------


class AutoCamExtractor(BaseExtractor):
    extractor_name = "auto_cam"
    schema_version = "1.0"

    async def extract(self, filename: str, body_bytes: bytes) -> ExtractionResult:
        try:
            wb = openpyxl.load_workbook(io.BytesIO(body_bytes), data_only=True)
        except Exception as exc:
            return ExtractionResult(
                status=ExtractionStatus.FAILED,
                schema_version=type(self).schema_version,
                data={},
                error_message=f"openpyxl failed to open {filename!r}: {exc}",
            )

        warnings: list[str] = []
        data: dict[str, Any] = {}

        # A "single-sheet CAM" variant (e.g. CAM_REPORT_<loan>.xlsx) only ships
        # one sheet by design; treating the three absent canonical sheets as
        # "missing" would mislabel the file as PARTIAL. We detect the variant
        # by workbook sheet count and omit the missing_sheet warnings + flag it
        # via data.variant for downstream consumers.
        is_single_sheet_variant = len(wb.sheetnames) == 1

        # Resolve sheet names (exact + fuzzy alias)
        sheet_map = _resolve_sheet(wb.sheetnames)

        for canonical in _CANONICAL_SHEETS:
            data_key = _SHEET_DATA_KEYS[canonical]
            actual = sheet_map.get(canonical)
            if actual is None:
                if not is_single_sheet_variant:
                    warnings.append(f"missing_sheet:{canonical}")
                data[data_key] = {}
                continue
            ws = wb[actual]
            sheet_data = _parse_sheet(
                ws,
                _SHEET_EXACT_MAPS[canonical],
                _SHEET_FUZZY_MAPS[canonical],
                right_fuzzy_list=_SHEET_RIGHT_COL_FUZZY.get(canonical),
            )
            data[data_key] = sheet_data

        if is_single_sheet_variant:
            data["variant"] = "single_sheet_cam"
            # If the sole sheet didn't match any canonical alias, flag as FAILED.
            if not sheet_map:
                return ExtractionResult(
                    status=ExtractionStatus.FAILED,
                    schema_version=type(self).schema_version,
                    data=data,
                    warnings=warnings,
                    error_message=(
                        f"Single-sheet workbook {filename!r} has no recognised CAM sheet"
                    ),
                )

        # Determine status — primary output is applicant_name + PAN, from any sheet.
        # Per-sheet key-field warnings (e.g. blank CIBIL) are informational only and
        # do not flip status; the spec only requires SUCCESS to mean the primary
        # identifiers are populated.
        missing_sheet_count = sum(1 for w in warnings if w.startswith("missing_sheet:"))
        if missing_sheet_count == len(_CANONICAL_SHEETS):
            return ExtractionResult(
                status=ExtractionStatus.FAILED,
                schema_version=type(self).schema_version,
                data=data,
                warnings=warnings,
                error_message="No expected sheets found in workbook.",
            )

        applicant_name = _find_in_data(data, ("applicant_name", "borrower_name"))
        pan = _find_in_data(data, ("pan", "pan_number"))
        if applicant_name is None:
            warnings.append("missing_applicant_name")
        if pan is None:
            warnings.append("missing_pan")

        # Single-sheet variant with primary extracted is SUCCESS (the file
        # delivered what it was designed to deliver). Multi-sheet variant
        # requires all 4 sheets present for SUCCESS.
        primary_present = bool(applicant_name and pan)
        if primary_present and (is_single_sheet_variant or missing_sheet_count == 0):
            status_val = ExtractionStatus.SUCCESS
        else:
            status_val = ExtractionStatus.PARTIAL

        return ExtractionResult(
            status=status_val,
            schema_version=type(self).schema_version,
            data=data,
            warnings=warnings,
        )
