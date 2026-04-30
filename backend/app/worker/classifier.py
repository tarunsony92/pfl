"""File classifier — decision tree for artifact subtype detection.

Spec §4.3 + plan §6: classify files by:
1. Filename regex patterns (case-insensitive)
2. Folder path heuristics (for images: BUSINESS_PREMISES, HOUSE_VISIT)
3. Content inspection (xlsx: CREDIT ASSESSMENT; html: equifax/cibil/etc)
4. Default to UNKNOWN

Signature: classify(filename: str, folder_path: str | None = None,
           body_bytes: bytes | None = None) -> ArtifactSubtype
"""

import re

from app.enums import ArtifactSubtype


def classify(
    filename: str,
    folder_path: str | None = None,
    body_bytes: bytes | None = None,
) -> ArtifactSubtype:
    """Classify a file to an ArtifactSubtype using filename patterns,
    folder hints, and content inspection.

    Args:
        filename: Base filename (e.g., "10006484_AADHAR_1.jpeg")
        folder_path: Full folder path (optional, used for image hints)
        body_bytes: Raw file content (optional, for xlsx/html inspection)

    Returns:
        ArtifactSubtype value
    """
    # Normalize filename to lowercase for pattern matching
    fn_lower = filename.lower()

    # Decision tree 1: Filename regex patterns (primary, most specific first)
    result = _classify_by_filename(fn_lower, filename)
    if result != ArtifactSubtype.UNKNOWN:
        return result

    # Decision tree 2: Folder path heuristics (images only)
    if folder_path and _is_image_extension(fn_lower):
        result = _classify_by_folder(folder_path)
        if result != ArtifactSubtype.UNKNOWN:
            return result

    # Decision tree 3: Content inspection (xlsx/html only)
    if body_bytes:
        ext = _get_extension(fn_lower)
        if ext in (".xlsx", ".xls"):
            result = _classify_xlsx_by_content(body_bytes)
            if result != ArtifactSubtype.UNKNOWN:
                return result
        elif ext == ".html":
            result = _classify_html_by_content(body_bytes)
            if result != ArtifactSubtype.UNKNOWN:
                return result

    return ArtifactSubtype.UNKNOWN


def _get_extension(fn_lower: str) -> str:
    """Extract file extension (with dot), lowercase."""
    if "." not in fn_lower:
        return ""
    return "." + fn_lower.split(".")[-1]


def _is_image_extension(fn_lower: str) -> bool:
    """Check if file is an image (jpg, jpeg, png, gif, bmp)."""
    ext = _get_extension(fn_lower)
    return ext in (".jpg", ".jpeg", ".png", ".gif", ".bmp")


def _classify_by_filename(fn_lower: str, filename_original: str) -> ArtifactSubtype:
    """Classify by filename regex patterns. Returns UNKNOWN if no match."""

    # CO_APPLICANT patterns — check first (most specific); allow any extension.
    # Seema-style: co_*, coapplicant_*. Ajay-style: coap_*, co-applicant-*.
    if re.search(r"(^|[_-])(co[_-]?applicant|coap|co)([_-])", fn_lower):
        if re.search(r"a{0,2}dha{1,2}r|\buid[_-]?(front|back)\b", fn_lower):
            return ArtifactSubtype.CO_APPLICANT_AADHAAR
        if re.search(r"(^|[_\s])pan([_\s.]|$)", fn_lower):
            return ArtifactSubtype.CO_APPLICANT_PAN

    # KYC patterns (image extensions only)
    # UID_FRONT / UID_BACK are common Aadhaar naming conventions.
    if _has_image_extension(fn_lower):
        if re.search(r"a{0,2}dha{1,2}r|\buid[_-]?(front|back)\b|^uid[_-]", fn_lower):
            return ArtifactSubtype.KYC_AADHAAR
        if re.search(r"(^|[_\s])pan([_\s.]|$)", fn_lower):
            return ArtifactSubtype.KYC_PAN
        if re.search(r"voter", fn_lower):
            return ArtifactSubtype.KYC_VOTER
        if re.search(r"(^|_)dl([_.]|$)|driving.{0,2}license", fn_lower):
            return ArtifactSubtype.KYC_DL
        if re.search(r"passport", fn_lower):
            return ArtifactSubtype.KYC_PASSPORT

    # RATION_CARD and ELECTRICITY_BILL (image or pdf)
    if re.search(r"ration", fn_lower) and _has_image_or_pdf_extension(fn_lower):
        return ArtifactSubtype.RATION_CARD
    if re.search(r"(electricity|elec).{0,2}bill", fn_lower) and _has_image_or_pdf_extension(
        fn_lower
    ):
        return ArtifactSubtype.ELECTRICITY_BILL

    # DEDUPE_REPORT (xlsx/xls only)
    # Finpage Customer-Dedupe export. Match before AUTO_CAM since both
    # consume .xlsx — DEDUPE is the more specific signal.
    if re.search(r"dedupe|duplicate", fn_lower) and _has_xlsx_extension(fn_lower):
        return ArtifactSubtype.DEDUPE_REPORT

    # AUTO_CAM (xlsx/xls only)
    # Seema-style: AUTO_CAM-*. Ajay-style: CAM_REPORT_*, bare loan-id .xlsx also
    # lands via content inspection below.
    if re.search(r"auto.{0,2}cam|cam[_-]?report", fn_lower) and _has_xlsx_extension(fn_lower):
        return ArtifactSubtype.AUTO_CAM

    # CHECKLIST (xlsx/xls only)
    if re.search(r"checklist", fn_lower) and _has_xlsx_extension(fn_lower):
        return ArtifactSubtype.CHECKLIST

    # PD_SHEET (docx only)
    if re.search(r"pd.{0,2}sheet", fn_lower) and fn_lower.endswith(".docx"):
        return ArtifactSubtype.PD_SHEET

    # Credit report HTML patterns (html only)
    if fn_lower.endswith(".html"):
        if re.search(r"equifax", fn_lower):
            return ArtifactSubtype.EQUIFAX_HTML
        if re.search(r"cibil", fn_lower):
            return ArtifactSubtype.CIBIL_HTML
        if re.search(r"highmark", fn_lower):
            return ArtifactSubtype.HIGHMARK_HTML
        if re.search(r"experian", fn_lower):
            return ArtifactSubtype.EXPERIAN_HTML

    # BANK_STATEMENT (pdf only)
    # Match "bank statement", "bank_statement", "bank_stmt" (Ajay-style).
    if re.search(r"bank.{0,2}(statement|stmt)", fn_lower) and fn_lower.endswith(".pdf"):
        return ArtifactSubtype.BANK_STATEMENT

    # KYC_VIDEO (mp4/mov only)
    if re.search(r"kyc.{0,2}video", fn_lower) and (
        fn_lower.endswith(".mp4") or fn_lower.endswith(".mov")
    ):
        return ArtifactSubtype.KYC_VIDEO

    # LOAN_AGREEMENT (pdf only)
    if re.search(r"loan.{0,2}agreement", fn_lower) and fn_lower.endswith(".pdf"):
        return ArtifactSubtype.LOAN_AGREEMENT

    # Loan docs (any extension) — LAPP, LAGR, DPN, NACH, KFS
    if re.search(r"(^|_)lapp(_|\.)", fn_lower):
        return ArtifactSubtype.LAPP
    if re.search(r"(^|_)lagr(_|\.)", fn_lower):
        return ArtifactSubtype.LAGR
    if re.search(r"(^|_)dpn(_|\.)", fn_lower):
        return ArtifactSubtype.DPN
    if re.search(r"(^|_)nach(_|\.)", fn_lower):
        return ArtifactSubtype.NACH
    if re.search(r"(^|_)kfs(_|\.)", fn_lower) and fn_lower.endswith(".pdf"):
        return ArtifactSubtype.KFS

    # UDYAM_REG (any extension)
    if re.search(r"udyam", fn_lower):
        return ArtifactSubtype.UDYAM_REG

    # PDC_CHEQUE — post-dated cheque given by the borrower as EMI security.
    # Match BEFORE BANK_ACCOUNT_PROOF so a "PDC.jpg" / "10006079_PDC_1.jpeg" /
    # "post_dated_cheque.png" doesn't fall through to the generic cheque
    # pattern below. We deliberately do NOT match a bare "cheque.jpg"
    # (which more often means cancelled cheque used for account proof).
    if re.search(
        r"(^|[_\-\s])pdc(_|\-|\s|\.)|post.{0,2}dated.{0,2}cheque|post.{0,2}dated.{0,2}chq|pdc.{0,2}cheque|pdc.{0,2}chq",
        fn_lower,
    ) and _has_image_or_pdf_extension(fn_lower):
        return ArtifactSubtype.PDC_CHEQUE

    # BANK_ACCOUNT_PROOF — cancelled cheque / passbook / bare "bank_account"
    # (image or pdf). The bare bank_account pattern matches Ajay-style
    # "10006079_BANK_ACCOUNT_1.PNG" — distinct from BANK_STATEMENT (pdf).
    if re.search(
        r"cancel{1,2}ed.{0,2}cheque|cancel{1,2}ed.{0,2}chq|passbook|bank.{0,2}account|bank.{0,2}acc.{0,2}proof|bank.{0,2}proof",
        fn_lower,
    ) and _has_image_or_pdf_extension(fn_lower):
        return ArtifactSubtype.BANK_ACCOUNT_PROOF

    # INCOME_PROOF — salary slip, ITR, Form 16 (image or pdf)
    if re.search(
        r"salary.{0,2}slip|salaryslip|itr[_.\s\d]|form.{0,2}16|income.{0,2}proof",
        fn_lower,
    ) and _has_image_or_pdf_extension(fn_lower):
        return ArtifactSubtype.INCOME_PROOF

    # TVR_AUDIO (mp3/wav/m4a) — TVR uploads are typically content-hashed
    # (e.g. f4dcb43843ca57f7870ca0e4f06d1a4n.mp3) so extension is the
    # reliable signal.
    if _has_audio_extension(fn_lower):
        return ArtifactSubtype.TVR_AUDIO

    return ArtifactSubtype.UNKNOWN


def _has_image_extension(fn_lower: str) -> bool:
    """Check if filename ends with image extension."""
    return fn_lower.endswith((".jpg", ".jpeg", ".png", ".gif", ".bmp"))


def _has_xlsx_extension(fn_lower: str) -> bool:
    """Check if filename ends with xlsx or xls."""
    return fn_lower.endswith((".xlsx", ".xls"))


def _has_audio_extension(fn_lower: str) -> bool:
    """Check if filename ends with audio extension (mp3/wav/m4a)."""
    return fn_lower.endswith((".mp3", ".wav", ".m4a"))


def _has_image_or_pdf_extension(fn_lower: str) -> bool:
    """Check if filename ends with image or pdf."""
    return fn_lower.endswith((".jpg", ".jpeg", ".png", ".gif", ".bmp", ".pdf"))


def _classify_by_folder(folder_path: str) -> ArtifactSubtype:
    """Classify by folder path hints (images only).

    Returns UNKNOWN if no match.
    """
    folder_upper = folder_path.upper()

    if "BUSINESS_PREMISES" in folder_upper:
        return ArtifactSubtype.BUSINESS_PREMISES_PHOTO
    if "HOUSE_VISIT" in folder_upper:
        return ArtifactSubtype.HOUSE_VISIT_PHOTO

    return ArtifactSubtype.UNKNOWN


def _classify_xlsx_by_content(body_bytes: bytes) -> ArtifactSubtype:
    """Classify xlsx by content inspection.

    xlsx is a zipped XML bundle, so literal-byte `in` checks don't work.
    We open the workbook with openpyxl and inspect sheet names, which is
    the cheapest reliable signal ("SystemCam" / "Elegibilty" / "CAM_REPORT"
    → AUTO_CAM; the fixture-shaped customer-dedupe workbook is excluded).

    Returns UNKNOWN on parse failure or when nothing recognisable matches.
    """
    try:
        import io

        import openpyxl

        wb = openpyxl.load_workbook(io.BytesIO(body_bytes), read_only=True, data_only=True)
    except Exception:
        return ArtifactSubtype.UNKNOWN

    sheet_names_lower = {(name or "").strip().lower() for name in wb.sheetnames}
    # Customer-dedupe workbook — Finpage identity-dedupe export.
    # Recognised by sheet name ('customer_dedupe' / 'customer dedupe').
    if any("customer_dedupe" in n or "customer dedupe" in n for n in sheet_names_lower):
        return ArtifactSubtype.DEDUPE_REPORT

    # AUTO_CAM sheet markers — full CAM has all four; single-sheet variants have
    # one of systemcam/cam_report/elegibilty.
    auto_cam_markers = {"systemcam", "cam_report", "elegibilty", "cm cam il", "health sheet"}
    if any(
        any(marker in name for marker in auto_cam_markers) for name in sheet_names_lower
    ):
        return ArtifactSubtype.AUTO_CAM

    # Also check the first worksheet's first row for "Credit Assessment" text —
    # covers older CAM exports that used a custom sheet name.
    try:
        first_ws = wb.worksheets[0]
        for row in first_ws.iter_rows(min_row=1, max_row=2, values_only=True):
            for cell in row:
                if cell and "credit assessment" in str(cell).lower():
                    return ArtifactSubtype.AUTO_CAM
    except Exception:
        pass

    return ArtifactSubtype.UNKNOWN


def _classify_html_by_content(body_bytes: bytes) -> ArtifactSubtype:
    """Classify html by content inspection.

    Returns UNKNOWN if no credit marker found.
    """
    body_lower = body_bytes.lower()

    # Credit report providers
    if b"equifax" in body_lower:
        return ArtifactSubtype.EQUIFAX_HTML
    if b"cibil" in body_lower:
        return ArtifactSubtype.CIBIL_HTML
    if b"highmark" in body_lower:
        return ArtifactSubtype.HIGHMARK_HTML
    if b"experian" in body_lower:
        return ArtifactSubtype.EXPERIAN_HTML

    return ArtifactSubtype.UNKNOWN
