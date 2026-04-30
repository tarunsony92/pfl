"""Dedupe snapshots HTTP endpoints."""

import io
import re
from datetime import UTC, datetime
from uuid import uuid4

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, UploadFile, status
from sqlalchemy import update
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import (
    get_current_user,
    get_session,
    get_storage_dep,
    require_role,
)
from app.config import get_settings
from app.enums import UserRole
from app.models.dedupe_snapshot import DedupeSnapshot
from app.models.user import User
from app.schemas.dedupe_snapshot import DedupeSnapshotRead
from app.services import audit
from app.services.storage import StorageService

router = APIRouter(prefix="/dedupe-snapshots", tags=["dedupe-snapshots"])


def _safe(name: str) -> str:
    """Sanitize filename for S3 key."""
    return re.sub(r"[^A-Za-z0-9._-]", "_", name)


@router.post("", response_model=DedupeSnapshotRead, status_code=status.HTTP_201_CREATED)
async def upload_dedupe_snapshot(
    file: UploadFile = File(...),
    actor: User = Depends(require_role(UserRole.ADMIN)),
    session: AsyncSession = Depends(get_session),
    storage: StorageService = Depends(get_storage_dep),
) -> DedupeSnapshotRead:
    """Upload a new dedupe snapshot (xlsx).

    - Admin-only
    - 50 MB max size
    - Deactivates previous active snapshot
    - Audit: dedupe_snapshot.uploaded + dedupe_snapshot.activated
    """
    if not file.filename:
        raise HTTPException(status_code=400, detail="filename required")

    # Read and check file size
    content = await file.read()
    max_bytes = get_settings().max_artifact_size_bytes
    if len(content) > max_bytes:
        raise HTTPException(
            status_code=status.HTTP_413_REQUEST_ENTITY_TOO_LARGE,
            detail=f"file too large (max {max_bytes} bytes)",
        )

    # Parse xlsx and count rows
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
    except Exception as exc:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid xlsx: {exc}",
        ) from exc

    # Get the Customer_Dedupe sheet or active sheet
    ws = wb["Customer_Dedupe"] if "Customer_Dedupe" in wb.sheetnames else wb.active

    # Row count = max_row - 1 (exclude header)
    row_count = max(ws.max_row - 1, 0) if ws.max_row else 0

    # Create snapshot model
    snapshot_uuid = uuid4()
    s3_key = f"dedupe/{snapshot_uuid}_{_safe(file.filename)}"
    snapshot = DedupeSnapshot(
        id=snapshot_uuid,
        uploaded_by=actor.id,
        uploaded_at=datetime.now(UTC),
        s3_key=s3_key,
        row_count=row_count,
        is_active=True,
    )

    # Upload to S3
    await storage.upload_object(
        s3_key,
        content,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    # Deactivate previous active snapshot (atomic via transaction)
    await session.execute(
        update(DedupeSnapshot).where(DedupeSnapshot.is_active.is_(True)).values(is_active=False)
    )

    # Insert new snapshot
    session.add(snapshot)
    await session.flush()

    # Audit logs
    await audit.log_action(
        session=session,
        actor_user_id=actor.id,
        action="dedupe_snapshot.uploaded",
        entity_type="dedupe_snapshot",
        entity_id=str(snapshot.id),
        before=None,
        after={"filename": file.filename, "row_count": row_count},
    )
    await audit.log_action(
        session=session,
        actor_user_id=actor.id,
        action="dedupe_snapshot.activated",
        entity_type="dedupe_snapshot",
        entity_id=str(snapshot.id),
        before=None,
        after={},
    )

    await session.commit()

    return DedupeSnapshotRead.model_validate(snapshot)


@router.get("", response_model=list[DedupeSnapshotRead])
async def list_dedupe_snapshots(
    actor: User = Depends(require_role(UserRole.ADMIN, UserRole.CEO, UserRole.CREDIT_HO)),
    session: AsyncSession = Depends(get_session),
    storage: StorageService = Depends(get_storage_dep),
) -> list[DedupeSnapshotRead]:
    """List all dedupe snapshots (ordered by uploaded_at DESC).

    Admin, CEO, or CREDIT_HO only.
    """
    from sqlalchemy import desc, select

    stmt = select(DedupeSnapshot).order_by(desc(DedupeSnapshot.uploaded_at))
    result = await session.execute(stmt)
    snapshots = result.scalars().all()

    # Attach download URLs (inline disposition so tools that fetch the URL
    # via fetch/Blob don't get bullied into a download by S3-stored
    # Content-Disposition metadata).
    reads = []
    for snapshot in snapshots:
        url = await storage.generate_presigned_download_url(
            snapshot.s3_key,
            expires_in=900,
            disposition="inline",
        )
        read = DedupeSnapshotRead.model_validate(snapshot).model_copy(update={"download_url": url})
        reads.append(read)

    return reads


@router.get("/active", response_model=DedupeSnapshotRead)
async def get_active_dedupe_snapshot(
    actor: User = Depends(get_current_user),
    session: AsyncSession = Depends(get_session),
    storage: StorageService = Depends(get_storage_dep),
) -> DedupeSnapshotRead:
    """Get the currently-active dedupe snapshot.

    Any authenticated user.
    Returns 404 if no active snapshot exists.
    """
    from sqlalchemy import select

    stmt = select(DedupeSnapshot).where(DedupeSnapshot.is_active.is_(True)).limit(1)
    result = await session.execute(stmt)
    snapshot = result.scalar_one_or_none()

    if snapshot is None:
        raise HTTPException(status_code=404, detail="No active dedupe snapshot")

    url = await storage.generate_presigned_download_url(
        snapshot.s3_key,
        expires_in=900,
        disposition="inline",
    )
    return DedupeSnapshotRead.model_validate(snapshot).model_copy(update={"download_url": url})
